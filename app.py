# ----- Librerias ------

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, Toplevel, Label, Button
import pandas as pd
import cv2  # Para la captura de QR
from pyzbar.pyzbar import decode  # Decodificar QR
import os
import requests
import json
from dotenv import load_dotenv
import urllib3
import re
import numpy as np
import threading
import queue
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Configuraciones -------------

# Deshabilitar las advertencias de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Cargar las variables del archivo .env
load_dotenv()

# Configuración de la API de GLPI
GLPI_URL = os.getenv("GLPI_URL")
USER_TOKEN = os.getenv("USER_TOKEN")
APP_TOKEN = os.getenv("APP_TOKEN")
PATH_EXCEL_ACTIVOS = os.getenv("PATH_EXCEL_ACTIVOS")
PATH_EXCEL_CONSUMIBLES = os.getenv("PATH_EXCEL_CONSUMIBLES")
IP_CAM_URL = os.getenv("IP_CAM_URL")


# Ruta del archivo Excel
ruta_excel = PATH_EXCEL_ACTIVOS

class GLPIApp:
    def __init__(self, root):
        # Crear el archivo Excel con las hojas "Computer", "Monitor" y "Consumables" si no existe
        self.crear_archivo_excel_con_hojas(ruta_excel, ["Computer", "Monitor", "Consumables"])
        self.root = root
        self.root.title("GLPI asset IT automation")
        self.root.geometry("800x600")
        self.style = ttk.Style()
        self.style.theme_use("clam")  # Puedes cambiar el tema a "clam", "alt", "default", "classic"
        self.configure_styles()
        self.create_widgets()
        self.obtener_todos_los_network_equipment_glpi(self.obtener_token_sesion())

    def salir(self):
        root.destroy()  

    ## --- GUI --- 
    
    def configure_styles(self):
        # Estilo del marco
        self.style.configure("TFrame", background="#E0F7FA")

        # Estilo de etiquetas
        self.style.configure("TLabel", background="#E0F7FA", foreground="#01579B", font=("Montserrat", 12))
        self.style.configure("Header.TLabel", background="#01579B", foreground="#FFFFFF", font=("Montserrat", 16, "bold"))

        # Estilo de botones con bordes redondeados y efectos suaves
        self.style.configure("Rounded.TButton",
                            background="#0288D1",
                            foreground="#FFFFFF",
                            font=("Roboto", 12),
                            padding=(15, 10),  # Aumentar el padding para suavizar
                            borderwidth=2,
                            relief="flat")  # 'flat' para quitar bordes bruscos

        # Aplicar efecto hover (cuando el mouse está sobre el botón)
        self.style.map("Rounded.TButton",
                    background=[("active", "#0277BD")],  # Cambio de color al pasar el mouse
                    relief=[("pressed", "groove")])  # Suaviza el clic en el botón

    def create_widgets(self):
        # Menú
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=file_menu)
        file_menu.add_command(label="Salir", command=self.root.quit)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ayuda", menu=help_menu)
        help_menu.add_command(label="Acerca de")

        # Pestañas
        tab_control = ttk.Notebook(self.root)
        tab_names = ["Registros Offline (Excel)", "Registros Online (GLPI)", "Excel -> GLPI (Sincronizacion Asincrona)"]
        frames = {}

        for name in tab_names:
            frames[name] = ttk.Frame(tab_control, padding="10")
            frames[name].pack(fill="both", expand=True)
            tab_control.add(frames[name], text=name)
            self.center_widgets(frames[name])  # Aplicar centrado global a cada pestaña

        tab_control.pack(expand=1, fill="both")

        # Laptops
        ttk.Label(frames["Registros Offline (Excel)"], text="Laptops", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Registros Offline (Excel)"], text="Escanear QR y registrar laptop (Dell/Mac)", command=self.registrar_laptop).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Registros Offline (Excel)"], text="Entregar laptop a un usuario", command=self.entregar_laptop).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Registros Offline (Excel)"], text="Salir", command=self.root.quit).grid(row=3, column=0, padx=10, pady=5)

        # Monitores
        ttk.Label(frames["Monitores"], text="Monitores", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Monitores"], text="Escanear QR y registrar monitores", command=self.manejar_qr_monitor).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Monitores"], text="Entregar monitor a un usuario", command=self.entregar_monitor).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Monitores"], text="Salir", command=self.root.quit).grid(row=3, column=0, padx=10, pady=5)

        # Consumibles
        ttk.Label(frames["Consumibles"], text="Consumibles", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Consumibles"], text="Agregar consumible", command=self.agregar_consumible).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Consumibles"], text="Quitar consumible", command=self.quitar_consumible).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Consumibles"], text="Salir", command=self.root.quit).grid(row=3, column=0, padx=10, pady=5)

        # Excel
        ttk.Label(frames["Excel/GLPI"], text="GLPI/Excel", style="Header.TLabel").grid(row=0, column=0, pady=10)
        #ttk.Button(frames["Excel/GLPI"], text="Registrar la última fila del Excel en GLPI", command=self.registrar_ultima_fila).grid(row=1, column=0, padx=10, pady=5)
        #ttk.Button(frames["Excel/GLPI"], text="Registrar un activo por nombre", command=self.registrar_por_nombre).grid(row=2, column=0, padx=10, pady=5)
        #ttk.Button(frames["Excel/GLPI"], text="Registrar todos los activos de Excel en GLPI", command=lambda: self.procesar_archivo_excel(ruta_excel)).grid(row=3, column=0, padx=10, pady=5)
        ttk.Button(frames["Excel/GLPI"], text="Warning: Extraer TODOS Datos de GLPI a Excel", command= self.extraer_datos_glpi_a_excel).grid(row=4, column=0, padx=10, pady=5)
        ttk.Button(frames["Excel/GLPI"], text="Sincronizar con GLPI", command=self.registrar_pendientes_glpi).grid(row=5, column=0, padx=10, pady=5)
        ttk.Button(frames["Excel/GLPI"], text="Salir", command=self.root.quit).grid(row=6, column=0, padx=10, pady=5)

    def center_widgets(self, frame):
        # Configurar la columna 0 del frame para centrar elementos
        frame.columnconfigure(0, weight=1)
        

    ## --- Excel ---

    def crear_archivo_excel_con_hojas(self, ruta, hojas):
        if not os.path.exists(ruta):
            wb = Workbook()
            for hoja in hojas:
                ws = wb.create_sheet(title=hoja)
                excel_headers = [
                    "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                    "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                    "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                    "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                    "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                    "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                    "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                    "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                    "invoice_number"
                ]
                ws.append(excel_headers)
            # Eliminar la hoja por defecto creada por Workbook
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(ruta)

    def crear_hoja_excel(self, wb, asset_type):
        if asset_type not in wb.sheetnames:
            ws = wb.create_sheet(title=asset_type)
            excel_headers = [
                "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                "invoice_number"
            ]
            ws.append(excel_headers)
        else:
            ws = wb[asset_type]
            excel_headers = [cell.value for cell in ws[1]]
        return ws, excel_headers

    ## --- Laptops ----

    def manejar_qr_laptop(self, flag):
        try: 
            if flag == "Register":
                #manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                manufacturer = self.menu_emergente_botones("Input", "Ingrese el fabricante del laptop (Dell/Mac):", "dell", "mac")
                serial_number = None
                
                if manufacturer in ["dell", "dell inc.", "dell inc", "dell inc.", "dell"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Dell Inc."
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Dell Inc."
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                elif manufacturer in ["mac", "mac inc.", "apple inc.", "apple"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{9}$', qr_data) or re.match(r'^S[A-Za-z0-9]{12}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Apple Inc"
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido o no corresponde al manufacturer. Intente nuevamente.")
                            return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Apple Inc"
                                    return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido no corresponde al manufacturer. Intente nuevamente.")
                            return
                else: 
                    messagebox.showerror("Error", "Fabricante no válido. Intente nuevamente.")
                    return
            elif flag == "Deliver":
                #manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                manufacturer = self.menu_emergente_botones("Input", "Ingrese el fabricante del laptop (Dell/Mac):", "dell", "mac")
                serial_number = None
                
                if manufacturer in ["dell", "dell inc.", "dell inc", "dell inc.", "dell"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Dell Inc."
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Dell Inc."
                                return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                    
                elif manufacturer in ["mac", "mac inc.", "apple inc.", "apple"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Apple Inc"
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Apple Inc"
                                return serial_number, manufacturer 
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return    
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                        
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
            return

    def registrar_laptop(self):
        try:
            messagebox.showinfo("Información", "--- Registrar Laptop en Excel ---")
            
            # Manejar el QR para obtener el serial y fabricante
            result = self.manejar_qr_laptop("Register")
            if result is None:
                messagebox.showerror("Error", "No se pudo obtener el serial number y el fabricante del laptop.")
                return
            
            serial_number, manufacturer = result
        
            # Determinar el nuevo nombre del laptop según el fabricante
            if manufacturer == "Dell Inc.":
                models_dict = self.buscar_modelos_latitude_precision(self.obtener_token_sesion())
                models_lista = list(models_dict.values())
                
                #model = simpledialog.askstring("Input", "Ingrese el modelo Dell (Latitude/Precision):").strip()
                model = self.menu_emergente_n_botones("Input Modelo Dell", "Ingrese el modelo Dell (Latitude/Precision):", models_lista)
                if model == "Latitude":
                    new_name = "None-Latitude"
                elif model == "Precision":
                    new_name = "None-Precision"
                else: 
                    messagebox.showerror("Error", "No se introdujo el modelo correctamente.")
            elif manufacturer == "Apple Inc": 
                model = "macbook"
                new_name = "None-MacBookPro"
            else:
                messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
                return
            
            location_name = simpledialog.askstring("Input", "Ingrese la ubicación del activo:").strip()

            # Crear diccionario con los datos del laptop
            asset_data = {
                "serial": serial_number,
                "manufacturers_id": manufacturer,
                "name": new_name,
                "status": "Stocked",  # Estado inicial en inventario
                "locations_id": location_name, # ID de ubicación obtenido de GLPI
                "computermodels_id": model
            }
            
            # Agregar al Excel usando la función modularizada
            self.agregar_a_excel(asset_data, "Computer")

            # Obtener sesión de GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            # Registrar en GLPI
            
            if self.verificar_existencia_asset(self, session_token, serial_number, asset_type="Computer") == False:
                # Solicitar ubicación y obtener location_id
                location_id = self.obtener_location_id(session_token, location_name)
                if not location_id:
                    messagebox.showerror("Error", f"No se encontró la ubicación '{location_name}' en GLPI.")
                    return
                
                # Obtener el ID del fabricante en GLPI
                manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer)
                if not manufacturer_id:
                    messagebox.showerror("Error", f"No se pudo encontrar el fabricante '{manufacturer}' en GLPI.")
                    return

                # Definir los encabezados HTTP 
                headers = {
                    "Content-Type": "application/json",
                    "Session-Token": session_token,
                    "App-Token": APP_TOKEN
                }

                # Preparar datos para la creación en GLPI
                payload = {
                    "input": {
                        "name": new_name,
                        "serial": serial_number,
                        "manufacturers_id": int(manufacturer_id),
                        "locations_id": int(location_id),  # Se agrega el location_id a GLPI
                        "status": "Stocked"
                    }
                }

                response = requests.post(f"{GLPI_URL}/Computer", headers=headers, json=payload, verify=False)

                if response.status_code == 201:
                    messagebox.showinfo("Éxito", f"Laptop con Service Tag '{serial_number}' registrada correctamente en GLPI.")
                else:
                    messagebox.showerror("Error", f"Error al registrar el laptop en GLPI: {response.status_code}")
                    try:
                        messagebox.showerror("Error", response.json())
                    except json.JSONDecodeError:
                        messagebox.showerror("Error", response.text)
            else:
                messagebox.showerror("Error", f"El laptop ya existe en el GLPI")

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")



if __name__ == "__main__":
    root = tk.Tk()
    app = GLPIApp(root)
    root.mainloop()