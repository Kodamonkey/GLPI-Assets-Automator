# ----- Librerias ------

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, Toplevel, Label, Button
import pandas as pd
import cv2  # Para la captura de QR
from pyzbar.pyzbar import decode  # Decodificar QR
import os
import requests
import json
from dotenv import load_dotenv
import urllib3
import re
import numpy as np
import threading
import queue
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Configuraciones -------------

# Deshabilitar las advertencias de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Cargar las variables del archivo .env
load_dotenv()

# Configuración de la API de GLPI
GLPI_URL = os.getenv("GLPI_URL")
USER_TOKEN = os.getenv("USER_TOKEN")
APP_TOKEN = os.getenv("APP_TOKEN")
PATH_EXCEL_ACTIVOS = os.getenv("PATH_EXCEL_ACTIVOS")
PATH_EXCEL_CONSUMIBLES = os.getenv("PATH_EXCEL_CONSUMIBLES")
IP_CAM_URL = os.getenv("IP_CAM_URL")

# Ruta del archivo Excel
#ruta_excel = PATH_EXCEL_ACTIVOS

# Obtener el directorio del script actual
current_dir = os.path.dirname(os.path.abspath(__file__))

# Ruta del archivo Excel
ruta_excel = os.path.join(current_dir, "activos.xlsx")

class GLPIApp:
    # ---- Configs. iniciales ----
    
    def __init__(self, root):
        # Crear el archivo Excel con las hojas "Computer", "Monitor" y "Consumables" si no existe
        self.crear_archivo_excel_con_hojas(ruta_excel, ["Computer", "Monitor", "Consumables", "Computer New", "Monitor New", "Consumables New"])
        self.root = root
        self.root.title("GLPI Asset Automator")
        self.root.geometry("1000x1000")
        self.style = ttk.Style()
        self.style.theme_use("clam")  # Puedes cambiar el tema a "clam", "alt", "default", "classic"
        self.configure_styles()
        self.create_widgets()
        self.obtener_id_consumible(self.obtener_token_sesion(), "Mochilas Kensington", "118")
        try:
            session_token = self.obtener_token_sesion()
            if session_token:
                 self.obtener_datos_glpi_a_excel(self.obtener_token_sesion())
        except Exception as e:
            print(f"Error al conectar con GLPI: {str(e)}")
            messagebox.showerror("Error", f"No se pudo conectar con GLPI, NO PODRAS REALIZAR SINCRONIZACIONES AUN: {str(e)}")

    def configure_styles(self):
        # Estilo del marco
        self.style.configure("TFrame", background="#E0F7FA")

        # Estilo de etiquetas
        self.style.configure("TLabel", background="#E0F7FA", foreground="#01579B", font=("Montserrat", 12))
        self.style.configure("Header.TLabel", background="#01579B", foreground="#FFFFFF", font=("Montserrat", 16, "bold"))

        # Estilo de botones con bordes redondeados y efectos suaves
        self.style.configure("Rounded.TButton",
                            background="#0288D1",
                            foreground="#FFFFFF",
                            font=("Roboto", 12),
                            padding=(15, 10),  # Aumentar el padding para suavizar
                            borderwidth=2,
                            relief="flat")  # 'flat' para quitar bordes bruscos

        # Aplicar efecto hover (cuando el mouse está sobre el botón)
        self.style.map("Rounded.TButton",
                    background=[("active", "#0277BD")],  # Cambio de color al pasar el mouse
                    relief=[("pressed", "groove")])  # Suaviza el clic en el botón

    def create_widgets(self):
        # Menú
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=file_menu)
        file_menu.add_command(label="Salir", command=self.root.quit)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ayuda", menu=help_menu)
        help_menu.add_command(label="Acerca de")

        # Pestañas
        tab_control = ttk.Notebook(self.root)
        tab_names = ["Laptops", "Monitores", "Consumibles", "Excel/GLPI"]
        frames = {}

        for name in tab_names:
            frames[name] = ttk.Frame(tab_control, padding="10")
            frames[name].pack(fill="both", expand=True)
            tab_control.add(frames[name], text=name)
            self.center_widgets(frames[name])  # Aplicar centrado global a cada pestaña

        tab_control.pack(expand=1, fill="both")

        # Laptops
        ttk.Label(frames["Laptops"], text="Laptops", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Laptops"], text="Registrar laptop offline (Excel)", command=lambda:self.registrar_laptop("Excel")).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Laptops"], text="Registrar laptop online (GLPI)", command=lambda:self.registrar_laptop("GLPI")).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Laptops"], text="Entregar laptop a un usuario", command=self.entregar_laptop).grid(row=3, column=0, padx=10, pady=5)
        ttk.Button(frames["Laptops"], text="Salir", command=self.root.quit).grid(row=4, column=0, padx=10, pady=5)

        # Monitores
        ttk.Label(frames["Monitores"], text="Monitores", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Monitores"], text="Registrar monitores offline (Excel)", command=lambda: self.manejar_qr_monitor("Excel")).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Monitores"], text="Registrar monitores online (GLPI)", command=lambda: self.manejar_qr_monitor("GLPI")).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Monitores"], text="Entregar monitor a un usuario", command=self.entregar_monitor).grid(row=3, column=0, padx=10, pady=5)
        ttk.Button(frames["Monitores"], text="Salir", command=self.root.quit).grid(row=4, column=0, padx=10, pady=5)

        # Consumibles
        ttk.Label(frames["Consumibles"], text="Consumibles", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Consumibles"], text="Agregar consumible offline (Excel)", command=lambda: self.agregar_consumible("Excel")).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Consumibles"], text="Agregar consumible online (GLPI)", command=lambda: self.agregar_consumible("GLPI")).grid(row=2, column=0, padx=10, pady=5)
        ttk.Button(frames["Consumibles"], text="Quitar consumible", command=self.quitar_consumible).grid(row=3, column=0, padx=10, pady=5)
        ttk.Button(frames["Consumibles"], text="Salir", command=self.root.quit).grid(row=4, column=0, padx=10, pady=5)

        # Excel
        ttk.Label(frames["Excel/GLPI"], text="GLPI/Excel", style="Header.TLabel").grid(row=0, column=0, pady=10)
        ttk.Button(frames["Excel/GLPI"], text="Sincronizar con GLPI", command=self.registrar_pendientes_glpi).grid(row=1, column=0, padx=10, pady=5)
        ttk.Button(frames["Excel/GLPI"], text="Salir", command=self.root.quit).grid(row=2, column=0, padx=10, pady=5)

    def center_widgets(self, frame):
        # Configurar la columna 0 del frame para centrar elementos
        frame.columnconfigure(0, weight=1)
        
    # ------- Funciones -----------

    # ------ Excel -------

    # Crear archivo Excel si no existe
    def crear_archivo_excel_con_hojas(self, ruta, hojas):
        if not os.path.exists(ruta):
            wb = Workbook()
            for hoja in hojas:
                ws = wb.create_sheet(title=hoja)
                excel_headers = [
                    "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                    "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                    "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                    "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                    "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                    "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                    "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                    "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                    "invoice_number"
                ]
                ws.append(excel_headers)
            # Eliminar la hoja por defecto creada por Workbook
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(ruta)

    def crear_hoja_excel(self, wb, asset_type):
        if asset_type not in wb.sheetnames:
            ws = wb.create_sheet(title=asset_type)
            excel_headers = [
                "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                "invoice_number"
            ]
            ws.append(excel_headers)
        else:
            ws = wb[asset_type]
            excel_headers = [cell.value for cell in ws[1]]
        return ws, excel_headers

    # ------ Tkinter menus --------
    def seleccionar_metodo_ingreso(self):
        """
        Muestra una ventana emergente con botones para elegir entre escanear o ingresar manualmente el Service Tag.
        """
        def set_opcion(opcion):
            self.metodo = opcion
            popup.destroy()

        self.metodo = None  # Variable para almacenar la opción del usuario
        popup = tk.Toplevel(self.root)
        popup.title("Seleccionar Método")
        popup.geometry("300x150")
        popup.transient(self.root)
        popup.grab_set()

        label = tk.Label(popup, text="¿Cómo desea ingresar el Service Tag?")
        label.pack(pady=10)

        boton_escanear = tk.Button(popup, text="Escanear", command=lambda: set_opcion("escanear"))
        boton_escanear.pack(pady=5)

        boton_manual = tk.Button(popup, text="Manual", command=lambda: set_opcion("manual"))
        boton_manual.pack(pady=5)

        popup.wait_window()  # Espera a que se cierre la ventana antes de continuar
        return self.metodo

    def custom_askyesnocancel(self, message):
        dialog = Toplevel(self.root)
        dialog.title("Confirmación")
        Label(dialog, text=message, padx=20, pady=20).pack()

        response = {"value": "cancel"}

        def set_response(value):
            response["value"] = value
            dialog.destroy()

        Button(dialog, text="Sí", command=lambda: set_response("yes")).pack(side="left", padx=10, pady=10)
        Button(dialog, text="No", command=lambda: set_response("no")).pack(side="left", padx=10, pady=10)
        Button(dialog, text="Sobrescribir todo", command=lambda: set_response("all")).pack(side="left", padx=10, pady=10)
        Button(dialog, text="Omitir todo", command=lambda: set_response("none")).pack(side="left", padx=10, pady=10)
        Button(dialog, text="Cancelar", command=lambda: set_response("cancel")).pack(side="left", padx=10, pady=10)

        dialog.transient(self.root)
        dialog.grab_set()
        self.root.wait_window(dialog)

        return response["value"]

    def salir(self):
        root.destroy()   

    def menu_emergente_botones(self, tittle, label, texto1, texto2):
        """
        Muestra una ventana emergente con botones para elegir entre escanear o ingresar manualmente el Service Tag.
        """
        def set_opcion(opcion):
            self.metodo = opcion
            popup.destroy()

        self.metodo = None  # Variable para almacenar la opción del usuario
        popup = tk.Toplevel(self.root)
        popup.title(f"{tittle}")
        popup.geometry("300x250")
        popup.transient(self.root)
        popup.grab_set()

        #label = tk.Label(popup, text="¿Cómo desea ingresar el Service Tag?")
        label = tk.Label(popup, text=f"{label}")
        label.pack(pady=10)

        boton_1 = tk.Button(popup, text=f"{texto1}", command=lambda: set_opcion(f"{texto1}"))
        boton_1.pack(pady=5)

        boton_2 = tk.Button(popup, text=f"{texto2}", command=lambda: set_opcion(f"{texto2}"))
        boton_2.pack(pady=5)

        popup.wait_window()  # Espera a que se cierre la ventana antes de continuar
        return self.metodo

    def menu_emergente_n_botones(self, title, label_text, botones):
        """
        Muestra una ventana emergente con una cantidad dinámica de botones.

        Parámetros:
        - title: Título de la ventana emergente.
        - label_text: Texto de la etiqueta en la ventana.
        - botones: Lista de nombres de los botones.

        Retorna:
        - La opción seleccionada por el usuario.
        """
        def set_opcion(opcion):
            self.opcion_seleccionada = opcion
            popup.destroy()

        self.opcion_seleccionada = None  # Variable para almacenar la opción del usuario
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.geometry("300x150")
        popup.transient(self.root)
        popup.grab_set()

        label = tk.Label(popup, text=label_text)
        label.pack(pady=10)

        for boton_texto in botones:
            boton = tk.Button(popup, text=boton_texto, command=lambda b=boton_texto: set_opcion(b))
            boton.pack(pady=5)

        popup.wait_window()  # Espera a que se cierre la ventana antes de continuar
        return self.opcion_seleccionada

    # ----- Excel y configs. ------
    def obtener_token_sesion(self):
        headers = {
            "Authorization": f"user_token {USER_TOKEN}",
            "App-Token": APP_TOKEN,
        }
        try:
            response = requests.get(f"{GLPI_URL}/initSession", headers=headers, verify=False)
            if response.status_code == 200:
                print("Sesión iniciada correctamente.")
                return response.json().get("session_token")
            else:
                print(f"Error al iniciar sesión: {response.status_code}")
                messagebox.showerror("Error", f"Error al iniciar sesión: {response.status_code}")
                return None
        except requests.exceptions.RequestException as e:
            print(f"Error al conectar con la API de GLPI: {e}")
            messagebox.showerror("Error", f"Error al conectar con la API de GLPI: {e}")
            return None
    
    def guardar_plantilla_txt(self, asset_data, nombre_archivo):
        with open(nombre_archivo, 'w') as file:
            for key, value in asset_data.items():
                file.write(f"{key}: {value}\n")
        messagebox.showinfo("Información", f"Plantilla guardada en {nombre_archivo}")

    def crear_hoja_excel(self, wb, asset_type):
        if asset_type not in wb.sheetnames:
            ws = wb.create_sheet(title=asset_type)
            excel_headers = [
            "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
            "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
            "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
            "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
            "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
            "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
            "status", "location", "department", "ip_address", "mac_address", "operating_system", 
            "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
            "invoice_number"
        ]
            ws.append(excel_headers)
        else:
            ws = wb[asset_type]
            excel_headers = [cell.value for cell in ws[1]]
        return ws, excel_headers

    def verificar_existencia_asset(self, session_token, serial_number, asset_type="Computer"):
        """
        Verifica si un activo con el número de serie proporcionado ya existe en GLPI.
        
        Parámetros:
        - session_token: Token de sesión de GLPI.
        - serial_number: Número de serie del activo a verificar.
        - asset_type: Tipo de activo (por defecto "Computer").
        
        Retorna:
        - True si el activo existe, False en caso contrario.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Mapeo de endpoints de diferentes tipos de activos
        endpoints = {
            "Computer": "Computer",
            "Monitor": "Monitor",
            "Network Equipment": "NetworkEquipment",
            "Consumables": "ConsumableItem",
            "Printer": "Printer",
            "Phone": "Phone",
            "Server": "Server",
            "Software": "Software"
        }

        # Obtiene el endpoint correspondiente al tipo de activo
        endpoint = endpoints.get(asset_type, "Computer")

        # Parámetros de búsqueda en GLPI
        params = {"searchText": serial_number, "range": "0-999"}

        try:
            response = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=params, verify=False)

            if response.status_code == 200:
                assets = response.json().get("data", [])

                # Verifica si el número de serie coincide con algún activo existente
                for asset in assets:
                    if asset.get('5', '').strip() == serial_number:
                        messagebox.showinfo("Información", f"El activo con número de serie '{serial_number}' ya existe en GLPI con ID: {asset.get('1')}, por ende no se añadira")
                        return True  # Activo encontrado

            else:
                messagebox.showerror("Error", f"Error al verificar existencia del activo: {response.status_code}")
                try:
                    messagebox.showerror("Error", response.json())
                except json.JSONDecodeError:
                    messagebox.showerror("Error", response.text)

        except requests.RequestException as e:
            messagebox.showerror("Error", f"Error en la solicitud HTTP: {str(e)}")

        return False  # Activo no encontrado

    def verificar_existencia_consumable_glpi(self, session_token, otherserial, asset_type="ConsumableItem"):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Mapeo de endpoints de diferentes tipos de activos
        endpoints = {
            "Computer": "Computer",
            "Monitor": "Monitor",
            "Network Equipment": "NetworkEquipment",
            "Consumables": "ConsumableItem",
            "Printer": "Printer",
            "Phone": "Phone",
            "Server": "Server",
            "Software": "Software"
        }

        # Obtiene el endpoint correspondiente al tipo de activo
        endpoint = endpoints.get(asset_type, "ConsumableItem")

        # Parámetros de búsqueda en GLPI
        params = {"searchText": otherserial, "range": "0-999"}

        try:
            response = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=params, verify=False)

            if response.status_code == 200:
                assets = response.json().get("data", [])

                # Verifica si el número de serie coincide con algún activo existente
                for asset in assets:
                    if asset.get('6', '').strip() == otherserial:
                        messagebox.showinfo("Información", f"El consumbile con numero de inventario '{otherserial}' ya existe en GLPI con ID: {asset.get('1')}, por ende no se añadira")
                        return True  # Activo encontrado

            else:
                messagebox.showerror("Error", f"Error al verificar existencia del activo: {response.status_code}")
                try:
                    messagebox.showerror("Error", response.json())
                except json.JSONDecodeError:
                    messagebox.showerror("Error", response.text)

        except requests.RequestException as e:
            messagebox.showerror("Error", f"Error en la solicitud HTTP: {str(e)}")

        return False  # Activo no encontrado

    def verificar_existencia_en_excel(self, serial_number):
        wb = load_workbook(ruta_excel)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                if row[4] == serial_number:  # Asumiendo que la columna "Serial Number" es la cuarta
                    messagebox.showinfo("Información", f"El activo con número de serie '{serial_number}' ya existe en la hoja '{sheet}' del Excel.")
                    return True
        return False
    
    def verificar_existencia_en_excel_consumible(self, serial_number):
        wb = load_workbook(ruta_excel)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                if row[5] == serial_number:  # Asumiendo que la columna "Serial Number" es la cuarta
                    messagebox.showinfo("Información", f"El activo con número de serie '{serial_number}' ya existe en la hoja '{sheet}' del Excel.")
                    return True
        return False

    def agregar_a_excel(self, asset_data, sheet_name):
        try:
            # Verificar que todas las columnas esperadas estén presentes en asset_data
            columnas_esperadas = [
                "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                "invoice_number"
            ]

            for columna in columnas_esperadas:
                if columna not in asset_data:
                    asset_data[columna] = ""  # Rellenar con cadena vacía si falta alguna columna

            # Cargar el archivo Excel existente
            wb = load_workbook(ruta_excel)
            
            # Verificar si la hoja existe, si no, crearla
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(columnas_esperadas)
            else:
                ws = wb[sheet_name]

            # Convertir la hoja a un DataFrame
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]  # Establecer la primera fila como encabezados
            df = df[1:]  # Eliminar la primera fila de encabezados

            #if self.verificar_existencia_en_excel(asset_data["serial"]):
            #    messagebox.showinfo("Información", f"El activo con serial '{asset_data['serial']}' ya está registrado en el Excel. No se agregará.")
            #    return

            # Agregar el nuevo registro al DataFrame
            nuevo_registro = pd.DataFrame([asset_data])
            df = pd.concat([df, nuevo_registro], ignore_index=True)

            # Escribir el DataFrame de vuelta a la hoja
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row):
                    ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)

            # Guardar el archivo Excel
            wb.save(ruta_excel)
            messagebox.showinfo("Éxito", "Datos registrados exitosamente en el Excel.")
            return

        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar los datos: {e}")

    def procesar_archivo_excel(self, ruta_archivo):
        try:
            df = pd.read_excel(ruta_archivo, skiprows=0)
            df.columns = df.columns.str.strip()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo Excel: {str(e)}")
            return

        columnas_necesarias = [
            "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
            "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
            "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
            "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
            "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
            "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
            "status", "location", "department", "ip_address", "mac_address", "operating_system", 
            "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
            "invoice_number"
        ]

        for columna in columnas_necesarias:
            if columna not in df.columns:
                messagebox.showerror("Error", f"La columna '{columna}' no existe en el archivo Excel.")
                return

        df = df.fillna("").astype(str)

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        # Mapeo para normalizar los tipos de assets
        asset_type_mapping = {
            "computer": "Computer",
            "network equipment": "Network Equipment",
            "consumables": "Consumables",
            "monitor": "Monitor"
        }

        for index, row in df.iterrows():
            try:
                asset_type = row["Asset Type"].strip().lower()
                asset_type = asset_type_mapping.get(asset_type, None)

                if not asset_type:
                    messagebox.showerror("Error", f"Tipo de asset desconocido: '{row['Asset Type']}' (Fila {index + 1}). Se omite esta fila.")
                    continue  # Saltar la fila si el tipo de asset no es válido

                # Obtener location_id
                location_id = self.obtener_location_id(session_token, row["location"].strip())
                if location_id is None:
                    messagebox.showerror("Error", f"No se pudo encontrar la ubicación: {row['location']} (Fila {index + 1})")
                    continue  # Saltar la fila si no se encuentra la ubicación

                # Obtener manufacturer_id
                manufacturer_id = self.obtener_manufacturer_id(session_token, row["manufacturers_id"].strip())
                if manufacturer_id is None:
                    messagebox.showerror("Error", f"No se pudo encontrar el fabricante: {row['manufacturers_id']} (Fila {index + 1})")
                    continue  # Saltar la fila si no se encuentra el fabricante

                asset_data = {
                    "name": row["name"].strip(),
                    "locations_id": location_id, 
                    "manufacturers_id": manufacturer_id,
                    "serial": row["serial"].strip(),
                    "comments": row["comment"].strip(),
                }

                messagebox.showinfo("Información", f"Procesando fila {index + 1}: {asset_data} como {asset_type}")
                self.registrar_asset(session_token, asset_data, asset_type)
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar la fila {index + 1}: {str(e)}")
                continue  # Continuar con la siguiente fila en caso de error

        messagebox.showinfo("Información", "Procesamiento del archivo Excel completado.")

    def limpiar_asset_data(self, asset_data):
        cleaned_data = {}
        for key, value in asset_data.items():
            if isinstance(value, float) and np.isnan(value):
                cleaned_data[key] = ""
            elif value is None:
                cleaned_data[key] = ""
            else:
                cleaned_data[key] = str(value).strip()
        return cleaned_data

    def registrar_asset(self, session_token, asset_data, asset_type):
        if self.verificar_existencia_asset(session_token, asset_data["serial"], asset_type):
            messagebox.showinfo("Información", f"El activo con número de serie {asset_data['serial']} ya existe en GLPI. No se realizará el registro.")
            return

        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        endpoint = {
            "Computer": "/Computer",
            "Monitor": "/Monitor",
            "Network Equipment": "/NetworkEquipment",
            "Consumables": "/ConsumableItem",
        }.get(asset_type, "/Computer")

        # Limpiar datos antes de enviarlos
        asset_data_clean = self.limpiar_asset_data(asset_data)

        # Crear la estructura correcta para la API de GLPI
        asset_data_array = {"input": [asset_data_clean]}
        
        response = requests.post(f"{GLPI_URL}{endpoint}", headers=headers, data=json.dumps(asset_data_array), verify=False)
        
        if response.status_code == 201:
            messagebox.showinfo("Éxito", f"Asset registrado exitosamente: {asset_data_clean['name']}")
        else:
            messagebox.showerror("Error", f"Error al registrar asset: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def registrar_ultima_fila(self, asset_type):
        df = pd.read_excel(ruta_excel, sheet_name=asset_type)
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        last_row = df.iloc[-1].to_dict()
        #messagebox.showinfo("Información", f"Última fila encontrada: {last_row}")

        # Reemplazar NaN con valores vacíos
        last_row = {key: ("" if pd.isna(value) else value) for key, value in last_row.items()}

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        if "name" not in last_row or "Asset Type" not in last_row:
            messagebox.showerror("Error", "La última fila no contiene las columnas esperadas.")
            return

        location_id = self.obtener_location_id(session_token, last_row["location"])
        if not location_id:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación: {last_row['location']}")
            return

        manufacturer_id = self.obtener_manufacturer_id(session_token, last_row["manufacturers_id"])
        if not manufacturer_id:
            messagebox.showerror("Error", f"No se pudo encontrar el fabricante: {last_row['manufacturers_id']}")
            return

        if location_id is None or manufacturer_id is None:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación o el fabricante para el activo '{last_row['name']}'")
            return

        # Preparar los datos para el registro en GLPI
        asset_data = {
            "name": last_row["name"].strip(),
            "locations_id": location_id,
            "manufacturers_id": manufacturer_id,
            "serial": last_row["serial"].strip(),
            "comments": last_row["comment"].strip() if last_row["comment"] else "N/A",
        }

        messagebox.showinfo("Información", f"Registrando asset: {asset_data}")

        self.registrar_asset(session_token, asset_data, last_row["Asset Type"])

    def registrar_por_nombre(self):
        df = pd.read_excel(ruta_excel)
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        nombre = simpledialog.askstring("Input", "Ingrese el nombre del activo a registrar:").strip()
        if not nombre:
            messagebox.showerror("Error", "No se ingresó un nombre válido.")
            return

        filtro = df[df["name"].str.lower() == nombre.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró el activo con el nombre '{nombre}' en el archivo Excel.")
            return

        row = filtro.iloc[0].to_dict()
        session_token = self.obtener_token_sesion()

        location_id = self.obtener_location_id(session_token, row["location"])
        manufacturer_id = self.obtener_manufacturer_id(session_token, row["manufacturers_id"])

        if location_id is None or manufacturer_id is None:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación o el fabricante para el activo '{row['name']}'")
            return

        # Preparar los datos para el registro en GLPI
        asset_data = {
            "name": row["name"].strip(),
            "locations_id": location_id,
            "manufacturers_id": manufacturer_id,
            "serial": row["serial"].strip(),
            "comments": row["comment"].strip(),
        }

        self.registrar_asset(session_token, asset_data, row["Asset Type"])

    def obtener_manufacturer_id(self, session_token, manufacturer_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        params = {'searchText': manufacturer_name, 'range': '0-999'}
        response = requests.get(f"{GLPI_URL}/Manufacturer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            manufacturers = response.json()
            for manufacturer in manufacturers:
                if manufacturer.get("name", "").strip().lower() == manufacturer_name.strip().lower():
                    return manufacturer["id"]
        return None

    def parse_qr_data_template(self, qr_string):
        asset_data = {}
        for line in qr_string.split("\n"):
            key, value = line.split(": ", 1)
            asset_data[key.strip()] = value.strip().replace('"', '')
        return asset_data

    def escanear_qr_con_celular(self):
        try:
            # Mostrar cuadro de diálogo de confirmación
            respuesta = messagebox.askokcancel("Confirmación", "¿Desea activar la cámara para escanear el QR?")
            if not respuesta:
                return None

            # Mostrar información adicional después de la confirmación
            messagebox.showinfo("Información", "Usando la cámara del celular. Presiona 'q' para salir.")

            def run_capture(result_queue):
                ip_cam_url = IP_CAM_URL  # Cambiar por la URL de la cámara IP
                camera_open = True

                try:
                    cap = cv2.VideoCapture(ip_cam_url)
                    if not cap.isOpened():
                        result_queue.put(("error", "No se pudo acceder a la cámara del celular. Reintentalo..."))
                        cap.release()
                        cv2.destroyAllWindows()
                        return

                    while camera_open:
                        ret, frame = cap.read()
                        if not ret:
                            result_queue.put(("error", "Error al obtener el cuadro de la cámara. Reintentando conexión..."))
                            break  # Sale del bucle interno para reintentar la conexión

                        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        qr_codes = decode(gray_frame)

                        for qr in qr_codes:
                            qr_data = qr.data.decode('utf-8')
                            flag = self.es_codigo_valido(qr_data)
                            if flag in ["dell", "mac", "monitor", "consumible"]:
                                result_queue.put(("info", f"Código QR {flag} escaneado: \n{qr_data}"))
                                result_queue.put(("data", qr_data))
                                camera_open = False
                                cap.release()
                                cv2.destroyAllWindows()
                                return
                            elif flag == "invalido":
                                # No mostrar mensaje repetitivo, solo continuar
                                continue

                        cv2.imshow("Escaneando QR con celular", frame)

                        if cv2.waitKey(1) & 0xFF == ord('q'):
                            camera_open = False
                            cap.release()
                            cv2.destroyAllWindows()
                            result_queue.put(("close", None))
                            return

                        # Verificar si la ventana fue cerrada
                        if cv2.getWindowProperty("Escaneando QR con celular", cv2.WND_PROP_VISIBLE) < 1:
                            camera_open = False
                            cap.release()
                            cv2.destroyAllWindows()
                            result_queue.put(("close", None))
                            return

                    cap.release()
                    cv2.destroyAllWindows()
                except Exception as e:
                    result_queue.put(("error", f"Se produjo un error inesperado: {str(e)}"))
                    cap.release()
                    cv2.destroyAllWindows()

            result_queue = queue.Queue()
            capture_thread = threading.Thread(target=run_capture, args=(result_queue,))
            capture_thread.start()

            while True:
                try:
                    msg_type, msg_content = result_queue.get(timeout=1)
                    if msg_type == "info":
                        messagebox.showinfo("Información", msg_content)
                    elif msg_type == "error":
                        messagebox.showerror("Error", msg_content)
                        return None
                    elif msg_type == "data":
                        return msg_content
                    elif msg_type == "close":
                        return None
                except queue.Empty:
                    continue
        except Exception as e:
            result_queue.put(("error", f"Se produjo un error inesperado: {str(e)}"))
            cv2.destroyAllWindows()

    def es_codigo_valido(self, qr_data):
        # lógica para determinar si un código QR es válido
        # verificar si el código QR coincide con ciertos patrones
        patrones_validos = [
            # Dell Service Tag: 7 caracteres alfanuméricos (excluyendo I, O y Q)
            r'^(?!.*[IOQ])[A-HJ-NP-Z0-9]{7}$',  # Ejemplo: 8B9X1R3

            # Dell Express Service Code (conversión numérica del Service Tag)
            r'^\d{10}$',  # Ejemplo: 1234567890
            
            # MacBook Serial Number: Inicia con C02 o FVX seguido de 8-10 caracteres alfanuméricos
            r'^C02[A-Z0-9]{8,10}$',  # Ejemplo: C02X3Y5VFH5
            r'^FVX[A-Z0-9]{8,10}$',  # Ejemplo: FVXJ45KLD9

            # MacBook Serial Numbers con prefijo opcional 'S'
            r'^S?(C02|FVX)[A-Z0-9]{8,10}$',  # Soporta opcional 'S' delante del serial

            # MacBook Air/Pro: Formato reciente con 12 caracteres alfanuméricos
            r'^[A-Z0-9]{12}$',  # Ejemplo: W8P6W5T5YV3C

            # Dell Monitor Service Tag: "CN" seguido de 10 caracteres alfanuméricos
            r'^CN[A-Z0-9]{10}$',  # Ejemplo: CN0V7X9J129025AN2AX1

            # Dell Monitor etiqueta más corta con prefijo opcional "S"
            r'^S?[A-Z0-9]{7}$',  # Ejemplo: SCN12345

            # Monitores Dell con prefijo "SN" o "S/N"
            r'^(SN|S/N)\s*[A-Z0-9]{7,12}$',  # Ejemplo: SN 5JH34X1
        ]

        patron_valido_dell = [       
            # Dell Service Tag: 7 caracteres alfanuméricos (excluyendo I, O y Q)
            r'^(?!.*[IOQ])[A-HJ-NP-Z0-9]{7}$',  # Ejemplo: 8B9X1R3

            # Dell Express Service Code (conversión numérica del Service Tag)
            r'^\d{10}$',  # Ejemplo: 1234567890   
        ]
        
        patron_valido_mac = [
            # MacBook Serial Number: Inicia con C02 o FVX seguido de 8-10 caracteres alfanuméricos
            r'^C02[A-Z0-9]{8,10}$',  # Ejemplo: C02X3Y5VFH5
            r'^FVX[A-Z0-9]{8,10}$',  # Ejemplo: FVXJ45KLD9

            # MacBook Serial Numbers con prefijo opcional 'S'
            r'^S?(C02|FVX)[A-Z0-9]{8,10}$',  # Soporta opcional 'S' delante del serial

            # MacBook Air/Pro: Formato reciente con 12 caracteres alfanuméricos
            r'^[A-Z0-9]{10}$',  # Ejemplo: W8P6W5T5YV3
            r'^[A-Z0-9]{11}$',  # Ejemplo: W8P6W5T5YV3C
        ]
        
        patron_valido_monitor = [
            # Dell Monitor Service Tag: "CN" seguido de 10 caracteres alfanuméricos
            r'^CN[A-Z0-9]{10}$',  # Ejemplo: CN0V7X9J129025AN2AX1

            # Dell Monitor etiqueta más corta con prefijo opcional "S"
            r'^S?[A-Z0-9]{7}$',  # Ejemplo: SCN12345

            # Monitores Dell con prefijo "SN" o "S/N"
            r'^(SN|S/N)\s*[A-Z0-9]{7,12}$',  # Ejemplo: SN 5JH34X1
        ]

        patron_valido_consumible = [
            # Consumibles: 6-12 caracteres alfanuméricos
            "Mochilas Kensington",
            "Mouse Logitech",
            "Teclado Dell",
        ]

        for patron in patron_valido_dell:
            if re.match(patron, qr_data):
                return "dell"
        
        for patron in patron_valido_mac:
            if re.match(patron, qr_data):
                return "mac"
        
        for patron in patron_valido_monitor:
            if re.match(patron, qr_data):
                return "monitor"
            
        for patron in patron_valido_consumible:
            if qr_data in patron_valido_consumible:
                return "consumible"
        
        return "invalido"

    def obtener_datos_glpi_a_excel(self, session_token):
        """
        Obtiene datos de "Computer", "Monitor" y "Consumables" en GLPI y los guarda en sus respectivas hojas del archivo Excel.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        endpoints = {
            "Computer": "Computer",
            "Monitor": "Monitor",
            "Consumables": "ConsumableItem"
        }
        
        for asset_type, endpoint in endpoints.items():
            response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params={"range": "0-999"}, verify=False)
            
            if response.status_code == 200:
                data = response.json()
                
                # Verificar si la respuesta es una lista o un diccionario
                data_list = data if isinstance(data, list) else data.get('data', [])
                
                if not data_list:
                    print(f"No se encontraron datos para {asset_type}.")
                    continue
                
                # Definir las columnas requeridas
                columnas = [
                    "id", "entities_id", "is_recursive", "name", "serial", "otherserial", "contact", "contact_num", 
                    "users_id_tech", "groups_id_tech", "date_mod", "comment", "locations_id", "networks_id", 
                    "networkequipmenttypes_id", "networkequipmentmodels_id", "manufacturers_id", "is_deleted", 
                    "is_template", "template_name", "users_id", "groups_id", "states_id", "ticket_tco", "is_dynamic", 
                    "uuid", "date_creation", "autoupdatesystems_id", "sysdescr", "cpu", "uptime", "last_inventory_update", 
                    "snmpcredentials_id", "links"
                ]
                
                # Convertir a DataFrame asegurando que solo se extraigan las columnas requeridas
                df = pd.DataFrame(data_list)
                columnas_presentes = [col for col in columnas if col in df.columns]
                df = df[columnas_presentes]  # Seleccionar solo las columnas especificadas
                
                # Verificar si el archivo ya existe
                if os.path.exists(ruta_excel):
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name=asset_type, index=False)
                else:
                    with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=asset_type, index=False)
                
                print(f"Datos de {asset_type} guardados en la hoja '{asset_type}' del archivo {ruta_excel}")
            else:
                print(f"Error: No se pudieron obtener los datos de {asset_type}. Código {response.status_code}")

    def extraer_datos_glpi_a_excel(self):
        try:
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            endpoints = {
                "Computer": "Computer",
                "Monitor": "Monitor",
                "Consumables": "ConsumableItem"
            }

            if os.path.exists(ruta_excel):
                wb = load_workbook(ruta_excel)
            else:
                wb = Workbook()

            for asset_type, endpoint in endpoints.items():
                all_data = []
                start = 0
                while True:
                    params = {
                        "range": f"{start}-{start + 999}"
                    }
                    response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)
                    if response.status_code == 200:
                        data = response.json()
                        if not data:
                            break
                        all_data.extend(data)
                        start += 1000
                    elif response.status_code == 400 and 'ERROR_RANGE_EXCEED_TOTAL' in response.text:
                        try:
                            error_message = response.json()
                            total_count_str = error_message[1].split(": ")[1].split(";")[0]
                            total_count = int(total_count_str)
                        except (IndexError, ValueError, KeyError) as e:
                            messagebox.showerror("Error", f"Error al procesar la respuesta de la API: {str(e)}")
                            return
                        if start >= total_count:
                            break
                        params = {
                            "range": f"{start}-{total_count - 1}"
                        }
                        response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)
                        if response.status_code == 200:
                            data = response.json()
                            all_data.extend(data)
                            break
                        else:
                            try:
                                error_message = response.json()
                            except json.JSONDecodeError:
                                error_message = response.text
                            messagebox.showerror("Error", f"Error al obtener datos de {asset_type}: {response.status_code}\n{error_message}")
                            return
                    else:
                        try:
                            error_message = response.json()
                        except json.JSONDecodeError:
                            error_message = response.text
                        messagebox.showerror("Error", f"Error al obtener datos de {asset_type}: {response.status_code}\n{error_message}")
                        return

                if not all_data:
                    messagebox.showinfo("Información", f"No se encontraron datos para {asset_type}.")
                    continue

                ws, excel_headers = self.crear_hoja_excel(wb, asset_type)

                existing_data = {row[3]: list(row) for row in ws.iter_rows(min_row=2, values_only=True) if len(row) > 3}

                # Crear un conjunto de números de serie obtenidos de GLPI
                glpi_serial_numbers = {item.get("serial", "") for item in all_data}

                # Eliminar filas en Excel que ya no existen en GLPI
                rows_to_delete = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    serial_number = row[3]
                    if serial_number not in glpi_serial_numbers:
                        rows_to_delete.append(row_idx)

                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)

                for item in all_data:
                    serial_number = item.get("serial", "")
                    name = item.get("name", "")
                    if serial_number in existing_data:
                        existing_row = existing_data[serial_number]
                        cambios = []
                        for idx, header in enumerate(excel_headers):
                            excel_value = existing_row[idx]
                            glpi_value = self.limpiar_valor(item.get(header, ""))
                            if excel_value != glpi_value:
                                cambios.append((header, excel_value, glpi_value))

                        if cambios:
                            respuesta = None
                            if len(cambios) == 1:
                                campo, valor_excel, valor_glpi = cambios[0]
                                respuesta = messagebox.askyesnocancel("Confirmación", f"El campo '{campo}' ha cambiado de '{valor_excel}' a '{valor_glpi}'. ¿Desea sobrescribir este cambio?")
                            else:
                                campos_cambiados = ", ".join([c[0] for c in cambios])
                                respuesta = messagebox.askyesnocancel("Confirmación", f"Los siguientes campos han cambiado: {campos_cambiados}. ¿Desea sobrescribir estos cambios?")
                            
                            if respuesta is None:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return  # Cancelar la operación
                            elif respuesta:
                                for idx, header in enumerate(excel_headers):
                                    if idx < len(existing_row):
                                        existing_row[idx] = self.limpiar_valor(item.get(header, ""))
                                # Actualizar la fila en la hoja de Excel
                                for col_idx, value in enumerate(existing_row, start=1):
                                    ws.cell(row=existing_row[0], column=col_idx, value=value)
                        else:
                            continue  # Omitir si no hay cambios
                    else:
                        row = [self.limpiar_valor(item.get(header, "")) for header in excel_headers]
                        ws.append(row)

            wb.save(ruta_excel)
            messagebox.showinfo("Información", "Datos extraídos y guardados en el archivo Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
    
    def limpiar_valor(self, valor):
        if isinstance(valor, list):
            return ", ".join(str(v) for v in valor)
        elif isinstance(valor, dict):
            return json.dumps(valor)
        return valor

    def actualizar_excel_al_iniciar(self):
        try:
            respuesta = messagebox.askyesno("Actualizar Excel", "¿Deseas actualizar el archivo Excel con los datos de GLPI?")
            if not respuesta:
                return

            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            endpoints = {
                "Computer": "Computer",
                "Monitor": "Monitor",
                "Consumables": "ConsumableItem"
            }

            if os.path.exists(ruta_excel):
                wb = load_workbook(ruta_excel)
            else:
                wb = Workbook()

            for asset_type, endpoint in endpoints.items():
                all_data = []
                start = 0
                while True:
                    params = {
                        "range": f"{start}-{start + 999}"
                    }
                    response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)
                    if response.status_code == 200:
                        data = response.json()
                        if not data:
                            break
                        all_data.extend(data)
                        start += 1000
                    elif response.status_code == 400 and 'ERROR_RANGE_EXCEED_TOTAL' in response.text:
                        try:
                            error_message = response.json()
                            total_count_str = error_message[1].split(": ")[1].split(";")[0]
                            total_count = int(total_count_str)
                        except (IndexError, ValueError, KeyError) as e:
                            messagebox.showerror("Error", f"Error al procesar la respuesta de la API: {str(e)}")
                            return
                        if start >= total_count:
                            break
                        params = {
                            "range": f"{start}-{total_count - 1}"
                        }
                        response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)
                        if response.status_code == 200:
                            data = response.json()
                            all_data.extend(data)
                            break
                        else:
                            try:
                                error_message = response.json()
                            except json.JSONDecodeError:
                                error_message = response.text
                            messagebox.showerror("Error", f"Error al obtener datos de {asset_type}: {response.status_code}\n{error_message}")
                            return
                    else:
                        try:
                            error_message = response.json()
                        except json.JSONDecodeError:
                            error_message = response.text
                        messagebox.showerror("Error", f"Error al obtener datos de {asset_type}: {response.status_code}\n{error_message}")
                        return

                if not all_data:
                    messagebox.showinfo("Información", f"No se encontraron datos para {asset_type}.")
                    continue

                ws, excel_headers = self.crear_hoja_excel(wb, asset_type)

                existing_data = {row[3]: list(row) for row in ws.iter_rows(min_row=2, values_only=True) if len(row) > 3}

                # Crear un conjunto de números de serie obtenidos de GLPI
                glpi_serial_numbers = {item.get("serial", "") for item in all_data}

                # Eliminar filas en Excel que ya no existen en GLPI y moverlas al final
                rows_to_move = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    serial_number = row[3]
                    if serial_number is None or serial_number not in glpi_serial_numbers:
                        rows_to_move.append(row)

                for row in rows_to_move:
                    ws.delete_rows(row[0])

                for item in all_data:
                    serial_number = item.get("serial", "")
                    if serial_number in existing_data:
                        existing_row = existing_data[serial_number]
                        for idx, header in enumerate(excel_headers):
                            if idx < len(existing_row):
                                existing_row[idx] = self.limpiar_valor(item.get(header, ""))
                        # Actualizar la fila en la hoja de Excel
                        for col_idx, value in enumerate(existing_row, start=1):
                            ws.cell(row=existing_row[0], column=col_idx, value=value)
                    else:
                        row = [self.limpiar_valor(item.get(header, "")) for header in excel_headers]
                        ws.append(row)

                # Añadir las filas movidas al final
                for row in rows_to_move:
                    ws.append([self.limpiar_valor(cell) for cell in row])

            wb.save(ruta_excel)
            messagebox.showinfo("Información", "Datos actualizados y guardados en el archivo Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def registrar_pendientes_glpi(self):
        try:
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión en GLPI.")
                return

            # Cargar el archivo Excel
            wb = load_workbook(ruta_excel)

            hojas_revisar = ["Computer New", "Monitor New", "Consumables New"]
            headers = {
                "Computer": "/Computer",
                "Monitor": "/Monitor",
                "Consumables": "/ConsumableItem"
            }

            for hoja in hojas_revisar:
                if hoja not in wb.sheetnames:
                    continue  # Si la hoja no existe, se ignora

                ws = wb[hoja]
                filas_a_eliminar = []  # Lista para almacenar los índices de las filas a eliminar
                tiene_datos = False  # Bandera para verificar si hay datos en el Excel

                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Extraer los datos de cada fila
                    asset_id = str(row[0]).strip() if row[0] else ""
                    name = str(row[3]).strip() if row[3] else ""
                    serial_number = str(row[4]).strip() if row[4] else ""
                    location_name = str(row[13]).strip() if row[13] else ""
                    manufacturer_name = str(row[19]).strip() if row[19] else ""
                    model = str(row[15]).strip() if row[15] else ""
                    stock_target = str(row[30]).strip() if row[30] else ""

                    if any([name, serial_number, location_name, manufacturer_name, model]):
                        tiene_datos = True

                    # Verificar si el dispositivo ya está registrado en GLPI
                    qr_info = {"name": name, "serial": serial_number}
                    existencia = self.verificar_existencia_asset(session_token, serial_number, hoja.split()[0])

                    if existencia:
                        # Si el dispositivo ya existe en GLPI, marcar la fila para eliminar
                        filas_a_eliminar.append(idx)
                        print(f"El dispositivo con nombre {name} y serial {serial_number} ya está registrado en GLPI. Se eliminará del Excel de dispositivos nuevos...")
                        messagebox.showinfo("Información", f"El dispositivo {name} ya está registrado en GLPI. Se eliminará del Excel de dispositivos nuevos...")
                    else:
                        # Obtener IDs necesarios
                        location_id = self.obtener_location_id(session_token, location_name)
                        manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer_name)
                        model_id = self.obtener_modelo_id(session_token, model)

                        # Estructura del dispositivo para registrar en GLPI
                        nuevo_dispositivo = {
                            "input": {
                                "name": name,
                                "serial": serial_number,
                                "locations_id": location_id,
                                "manufacturers_id": manufacturer_id,
                                "computermodels_id": model_id,
                                "status": "Stocked"
                            }
                        }

                        if hoja == "Consumables New" and stock_target:
                            nuevo_dispositivo["input"]["stock_target"] = stock_target

                        try:
                            # Hacer la solicitud POST a la API de GLPI para registrar el nuevo dispositivo
                            response = requests.post(f'{GLPI_URL}{headers[hoja.split()[0]]}', headers=headers, json=nuevo_dispositivo, verify=False)

                            # Verificar si la respuesta fue exitosa (código 200 o 201)
                            if response.status_code in [200, 201]:
                                created_device = response.json()
                                print(f"Dispositivo registrado en GLPI: {json.dumps(created_device, indent=4)}")
                                messagebox.showinfo("Información", f"Dispositivo {name} agregado correctamente a GLPI desde el Excel de dispositivos nuevos...")
                            else:
                                print(f"Error al registrar el dispositivo en GLPI: {response.status_code} - {response.text}")
                                messagebox.showerror("Error", f"No se pudo agregar el dispositivo {name} a GLPI.")
                        except requests.exceptions.RequestException as e:
                            print(f"Error al conectar con la API de GLPI: {e}")
                            messagebox.showerror("Error", f"Error de conexión con la API de GLPI para el dispositivo {name}.")

                # Eliminar filas en orden inverso para evitar problemas de desplazamiento de índices
                for i in sorted(filas_a_eliminar, reverse=True):
                    ws.delete_rows(i)

            # Guardar el archivo Excel con los cambios
            wb.save(ruta_excel)

            # Verificar si el Excel quedó vacío
            if not tiene_datos or ws.max_row == 1:
                print("El Excel de dispositivos nuevos está vacío desde antes o después de la eliminación.")
                messagebox.showwarning("Advertencia", "El Excel de dispositivos nuevos está vacío después de la eliminación.")

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def buscar_asset_en_glpi(self, session_token, serial_number, name, asset_type):
        """
        Busca un activo en GLPI por serial_number o name.
        Retorna el ID si existe, o None si no se encuentra.
        """
        try:
            # Definir los endpoints de búsqueda según el tipo de activo
            search_endpoints = {
                
                "Computer": "search/Computer",
                "Monitor": "search/Monitor",
                "Consumables": "search/ConsumableItem",
            }

            endpoint = search_endpoints.get(asset_type)
            if not endpoint:
                messagebox.showerror("Error", f"Tipo de activo '{asset_type}' no válido para búsqueda en GLPI.")
                return None

            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            # Intentar buscar por número de serie si está disponible
            if serial_number:
                params = {"searchText": serial_number, "range": "0-10"}
                response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)

                if response.status_code == 200:
                    assets = response.json().get("data", [])
                    if assets:
                        return assets[0].get("1")  # Retorna el ID del primer resultado encontrado

            # Intentar buscar por nombre si no se encontró por serial
            if name:
                params = {"searchText": name, "range": "0-10"}
                response = requests.get(f"{GLPI_URL}/{endpoint}", headers=headers, params=params, verify=False)

                if response.status_code == 200:
                    assets = response.json().get("data", [])
                    if assets:
                        return assets[0].get("1")  # Retorna el ID del primer resultado encontrado

        except Exception as e:
            messagebox.showerror("Error", f"Error al buscar asset en GLPI: {str(e)}")

        return None  # Si no se encontró el activo, retorna None

    def obtener_asset_glpi(self, session_token, asset_id, asset_type):
        """
        Obtiene la información actual de un activo en GLPI.
        """
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        endpoints = {
            "Computer": "/Computer",
            "Monitor": "/Monitor",
            "Consumables": "/ConsumableItem"
        }

        endpoint = endpoints.get(asset_type, "/Computer")

        response = requests.get(f"{GLPI_URL}{endpoint}/{asset_id}", headers=headers, verify=False)

        if response.status_code == 200:
            return response.json()
        else:
            return None

    # ---- Consumibles ------
    def actualizar_excel_consumible(self, nombre, inventory_number, location, stock):
        if not os.path.exists(ruta_excel):
            self.crear_archivo_excel_con_hojas(ruta_excel, ["Consumables"])

        wb = load_workbook(ruta_excel)
        ws, excel_headers = self.crear_hoja_excel(wb, "Consumables")

        # Convertir a string y manejar NaN para evitar errores
        inventory_number_str = str(inventory_number).strip().lower()
        location_str = str(location).strip() if pd.notna(location) else ""

        session_token = self.obtener_token_sesion()
        location_id = self.obtener_location_id(session_token, location_str)

        if location_id is None:
            messagebox.showerror("Error", "No se pudo obtener el ID de la ubicación. No se registrará el consumible.")
            return

        # Verificar si el consumible ya está registrado
        for row in ws.iter_rows(min_row=2, values_only=False):
            if (row[excel_headers.index("name")].value or "").strip().lower() == nombre.lower() and \
            (row[excel_headers.index("otherserial")].value or "").strip().lower() == inventory_number_str:
                row[excel_headers.index("stock_target")].value = stock
                wb.save(ruta_excel)
                messagebox.showinfo("Información", f"El stock del consumible '{nombre}' ha sido actualizado en el Excel.")
                return

        # Crear un diccionario con los datos a agregar
        nuevo_consumible = {col: None for col in excel_headers}  # Inicializar con None en todas las columnas
        nuevo_consumible["name"] = nombre
        nuevo_consumible["otherserial"] = inventory_number_str
        nuevo_consumible["locations_id"] = location_id
        nuevo_consumible["stock_target"] = stock

        # Ordenar los valores para alinearlos con las columnas del Excel
        nuevo_consumible_fila = [nuevo_consumible[col] for col in excel_headers]

        # Agregar la nueva fila correctamente alineada
        ws.append(nuevo_consumible_fila)

        # Guardar el archivo
        wb.save(ruta_excel)
        messagebox.showinfo("Información", f"El consumible '{nombre}' ha sido registrado en el Excel correctamente.")
    
    def agregar_consumible_en_excel(self, nombre_consumible, inventory_number, location, cantidad):
        try:
            # Cargar Excel y verificar si el consumible ya existe
            df = pd.read_excel(ruta_excel, sheet_name="Consumables")
            df.columns = df.columns.str.strip()  # Eliminar espacios en los nombres de columnas

            filtro = df[df["otherserial"].astype(str).str.lower() == inventory_number.lower()]
            if not filtro.empty:
                stock_actual = filtro.iloc[0]["stock_target"]
                nuevo_stock = stock_actual + cantidad
                messagebox.showinfo("Información", f"Consumible '{nombre_consumible}' encontrado en el Excel. Actualizando stock a {nuevo_stock}.")
            else:
                nuevo_stock = cantidad
                messagebox.showinfo("Información", f"No se encontró el consumible '{nombre_consumible}' en el Excel. Creando nuevo con stock {nuevo_stock}.")

            # Registrar en Excel
            self.actualizar_excel_consumible(nombre_consumible, inventory_number, location, nuevo_stock)
            messagebox.showinfo("Éxito", "Consumible registrado exitosamente en el Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar el consumible en Excel: {str(e)}")

    def agregar_consumible_en_glpi(self, nombre_consumible, inventory_number, location, cantidad):
        try:
            # Obtener sesión de GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            # Obtener o crear el consumible en GLPI
            consumible_id = self.obtener_id_consumible(session_token, nombre_consumible, inventory_number)
            if not consumible_id:
                messagebox.showinfo("Información", f"No se encontró el consumible '{nombre_consumible}' en GLPI. Creando uno nuevo...")
                consumible_id = self.crear_consumible(session_token, nombre_consumible, inventory_number, location, cantidad)
                if not consumible_id:
                    messagebox.showerror("Error", "Error al crear el consumible en GLPI.")
                    return

            # Actualizar stock en GLPI
            stock_actual = self.obtener_stock_actual(session_token, consumible_id)
            nuevo_stock = stock_actual + cantidad
            self.actualizar_stock_glpi(session_token, consumible_id, nuevo_stock)
            messagebox.showinfo("Información", f"Consumible '{nombre_consumible}' actualizado a {nuevo_stock} unidades en GLPI.")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def agregar_consumible(self, type_register):
        """
        Maneja la adición de consumibles al stock, validando si ya existen en Excel y GLPI.
        Si no existen, los crea antes de actualizar el stock.
        """
        try:
            messagebox.showinfo("Información", "--- Agregar Consumible al Stock ---")

            # Obtener número de inventario (QR o manual)
            #inventory_number = self.obtener_numero_inventario()
            nombre_consumible = self.obtener_numero_inventario()
            #if not inventory_number:
            #    return  # Error ya manejado en `obtener_numero_inventario`

            #nombre_consumible = simpledialog.askstring("Input", "Ingrese el nombre del consumible: ").strip()
            location = simpledialog.askstring("Input", "Ingrese la ubicación del consumible: ").strip()
            cantidad = int(simpledialog.askstring("Input", "Ingrese la cantidad a agregar al stock: "))


            if type_register == "Excel":
                # Registrar en Excel
                #self.agregar_consumible_en_excel(nombre_consumible, inventory_number, location, cantidad)
                self.agregar_consumible_en_excel(nombre_consumible, nombre_consumible, location, cantidad)
            elif type_register == "GLPI":
                # Registrar en GLPI
                #self.agregar_consumible_en_glpi(nombre_consumible, inventory_number, location, cantidad)
                self.agregar_consumible_en_glpi(nombre_consumible, nombre_consumible, location, cantidad)

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def obtener_numero_inventario(self):
        """
        Obtiene el número de inventario del consumible, ya sea mediante escaneo QR o entrada manual.
        Devuelve el número si es válido, o None si el usuario cancela o hay un error.
        """
        try:
            metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar manualmente? (qr/manual): ").strip().lower()
            if metodo == "qr":
                qr_data = self.escanear_qr_con_celular()
            elif metodo == "manual":
                qr_data = simpledialog.askstring("Input", "Ingrese el número de inventario o activo: ").strip()
            else:
                messagebox.showerror("Error", f"Método no válido: {metodo}")
                return None

            if not qr_data:
                messagebox.showerror("Error", "No se detectó ningún código QR o número de inventario inválido.")
                return None

            inventory_number = qr_data.strip()
            messagebox.showinfo("Información", f"Inventory Number detectado: {inventory_number}")

            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este número de inventario '{inventory_number}'? (sí/no): ").strip().lower()
            #if confirmacion not in ["sí", "si"]:
            #    messagebox.showinfo("Información", "Operación cancelada por el usuario.")
            #    return None

            return inventory_number

        except Exception as e:
            messagebox.showerror("Error", f"Error al obtener el número de inventario: {str(e)}")
            return None

    def actualizar_stock_glpi(self, session_token, consumible_id, nuevo_stock):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        payload = {
            "input": {
                "id": consumible_id,
                "stock_target": nuevo_stock
            }
        }

        response = requests.put(f"{GLPI_URL}/ConsumableItem/{consumible_id}", headers=headers, json=payload, verify=False)

        if response.status_code != 200:
            #messagebox.showinfo("Información", f"Stock del consumible ID {consumible_id} actualizado correctamente a {nuevo_stock}.")
            messagebox.showerror("Error", f"Error al actualizar el stock en GLPI: {response.status_code}")
            try:
                print(f"Error {response.json()}")
            except json.JSONDecodeError:
                print(f"Error {response.text}")

    def crear_consumible(self, session_token, nombre, inventory_number, location, stock_target):
        """
        Crea un nuevo consumible en GLPI con la información proporcionada.
        """

        try:
            # Manejo seguro de NaN y conversión a string
            location = "" if pd.isna(location) else str(location).strip()
            inventory_number = "" if pd.isna(inventory_number) else str(inventory_number).strip()

            # Si la ubicación sigue vacía, solicitarla manualmente al usuario
            if not location:
                location = simpledialog.askstring("Input", "Ingrese la ubicación del consumible:").strip()

            # Si la ubicación sigue vacía después de solicitarla, detener el proceso
            if not location:
                messagebox.showerror("Error", "Ubicación no proporcionada. No se puede registrar el consumible.")
                return None

            # Obtener location_id
            location_id = self.obtener_location_id(session_token, location)
            if not location_id:
                messagebox.showerror("Error", f"No se encontró la ubicación '{location}' en GLPI.")
                return None

            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            payload = {
                "input": {
                    "name": nombre,
                    "otherserial": inventory_number,
                    "locations_id": int(location_id),
                    "stock_target": stock_target
                }
            }

            response = requests.post(f"{GLPI_URL}/ConsumableItem", headers=headers, json=payload, verify=False)

            if response.status_code == 201:
                consumible_id = response.json().get("id")
                #messagebox.showinfo("Información", f"Consumible '{nombre}' creado exitosamente con ID {consumible_id}.")
                return consumible_id
            else:
                messagebox.showerror("Error", f"Error al crear el consumible en GLPI: {response.status_code}")
                try:
                    print(f"Error: {response.json()}")
                except json.JSONDecodeError:
                    messagebox.showerror(f"Error: {response.text}")
                return None

        except Exception as e:
            messagebox.showerror("Error", f"Error al crear el consumible en GLPI: {str(e)}")
            return None

    def quitar_consumible(self):
        messagebox.showinfo("Información", "--- Quitar Consumible del Stock ---")

        metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar manualmente? (qr/manual): ").strip().lower()

        if metodo == "qr":
            qr_data = self.escanear_qr_con_celular()
            if qr_data:
                inventory_number = qr_data.strip()
                messagebox.showinfo("Información", f"Inventory Number detectado: {inventory_number}")
            else:
                messagebox.showerror("Error", "No se detectó ningún código QR.")
                return
        else:
            inventory_number = simpledialog.askstring("Input", "Ingrese el número de inventario o activo: ").strip()

        df = pd.read_excel(ruta_excel, sheet_name="Consumables")
        df.columns = df.columns.str.strip()  # Asegurar que no haya espacios en los nombres de columnas

        # Buscar el consumible en el Excel por inventory_number
        filtro = df[df["otherserial"].astype(str).str.lower() == inventory_number.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró el consumible con Inventory Number '{inventory_number}' en el archivo Excel.")
            return

        # Obtener datos existentes del consumible
        nombre_consumible = filtro.iloc[0]["name"]
        location = filtro.iloc[0]["location"]
        stock_actual = int(filtro.iloc[0]["stock_target"])

        cantidad = int(simpledialog.askstring("Input", f"Ingrese la cantidad a retirar (Stock actual: {stock_actual}): "))

        if stock_actual < cantidad:
            messagebox.showerror("Error", "No se puede retirar más cantidad de la que hay en stock.")
            return

        nuevo_stock = stock_actual - cantidad

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        consumible_id = self.obtener_id_consumible(session_token, nombre_consumible, inventory_number)
        if not consumible_id:
            messagebox.showerror("Error", f"No se encontró el consumible '{nombre_consumible}' en GLPI.")
            return

        self.actualizar_stock_glpi(session_token, consumible_id, nuevo_stock)
        messagebox.showinfo("Información", f"Consumible '{nombre_consumible}' actualizado a {nuevo_stock} unidades en stock.")

        # Actualizar en Excel
        self.actualizar_excel_consumible(nombre_consumible, inventory_number, location, nuevo_stock)

    def obtener_id_consumible(self, session_token, nombre_consumible, inventory_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": nombre_consumible.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/ConsumableItem", headers=headers, params=params, verify=False)
        
        if response.status_code == 200:
            consumibles = response.json()
            print(f"Consumibles encontrados: {json.dumps(consumibles, indent=4)}")
            for consumible in consumibles:
                # Convertir a cadena de texto y limpiar espacios en blanco
                consumible_name = str(consumible.get("name", "")).strip().lower()
                consumible_serial = str(consumible.get("otherserial", "")).strip().lower()
                inventory_number_str = str(inventory_number).strip().lower()

                if consumible_name == nombre_consumible.strip().lower() and consumible_serial == inventory_number_str:
                    return consumible["id"]

        messagebox.showinfo("Información", f"No se encontró el consumible con nombre '{nombre_consumible}' y número de inventario '{inventory_number}' en GLPI.")
        return None

    def obtener_stock_actual(self, session_token, consumible_id):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        response = requests.get(f"{GLPI_URL}/ConsumableItem/{consumible_id}", headers=headers, verify=False)

        if response.status_code == 200:
            return int(response.json().get("stock_target", 0))
        else:
            messagebox.showerror("Error", f"Error al obtener el stock del consumible ID {consumible_id}: {response.status_code}")
            return 0

    # ------ Monitor --------
    def manejar_qr_monitor(self, type_register):
        #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el número de serie manualmente? (qr/manual): ").strip().lower()
        metodo = self.seleccionar_metodo_ingreso()

        if metodo == "escanear":
            qr_data = self.escanear_qr_con_celular()
            if qr_data:
                if re.match(r'^CN[A-Z0-9]{10}$', qr_data) or re.match(r'^S?[A-Z0-9]{7}$', qr_data) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', qr_data):
                    messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")

                    serial_number = qr_data.strip()

                    if serial_number:
                        #messagebox.showinfo("Información", f"Serial Number detectado: {serial_number}")
                        #confirmacion = simpledialog.askstring("Confirmación", "¿Es correcto este Serial Number, desea continuar? (sí/no): ").strip().lower()
                        #confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number {serial_number}, desea continuar?:", "Si", "No")
                        
                        #if confirmacion not in ["sí", "si", "Si", "Sí"]:
                        #    messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                        #    return
                        #else:
                        if self.verificar_existencia_en_excel(serial_number):
                            #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                            return
                        #asset_data = self.procesar_qr_monitor(serial_number)
                        #self.agregar_a_excel(asset_data, "Monitor")
                        self.registrar_monitor(qr_data)
                        #self.subir_monitor_glpi(asset_data)
                else:
                    messagebox.showerror("Error", "Código QR no corresponde a un monitor.")
            else:
                messagebox.showerror("Error", "No se detectó ningún código QR.")
        elif metodo == "manual":
            serial_number = simpledialog.askstring("Input", "Ingrese el número de serie del monitor: ").strip()
            if re.match(r'^CN[A-Z0-9]{10}$', serial_number) or re.match(r'^S?[A-Z0-9]{7}$', serial_number) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', serial_number):
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")

                if serial_number:
                    messagebox.showinfo("Información", f"Serial Number detectado: {serial_number}")
                    #confirmacion = simpledialog.askstring("Confirmación", "¿Es correcto este Serial Number, desea continuar? (sí/no): ").strip().lower()
                    confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number {serial_number}, desea continuar?:", "Si", "No")
                    if confirmacion not in ["sí", "si", "Si", "Sí"]:
                        messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                        return
                    else:
                        if self.verificar_existencia_en_excel(serial_number):
                            #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                            return
                        else:
                            self.registrar_monitor(qr_data)
            else:
                messagebox.showerror("Error", "El número de serie ingresado no corresponde a un monitor válido.")
        else:
            messagebox.showerror("Error", "Método no válido. Intente nuevamente.")

    def subir_monitor_glpi(self, asset_data):
        """
        Sube el monitor registrado en Excel a GLPI.
        """
        try:
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return
            if self.verificar_existencia_asset(session_token, asset_data["serial"], asset_type="Monitor") == False:
                
                manufacturer_id = self.obtener_manufacturer_id(session_token, asset_data["manufacturers_id"])
                if not manufacturer_id:
                    messagebox.showerror("Error", f"No se pudo encontrar el fabricante '{asset_data['manufacturers_id']}' en GLPI.")
                    return

                headers = {
                    "Content-Type": "application/json",
                    "Session-Token": session_token,
                    "App-Token": APP_TOKEN
                }

                payload = {
                    "input": {
                        "name": asset_data["name"],
                        "serial": asset_data["serial"],
                        "manufacturers_id": int(manufacturer_id),
                        "locations_id": int(asset_data["locations_id"]),
                        "status": "Stocked"
                    }
                }

                response = requests.post(f"{GLPI_URL}/Monitor", headers=headers, json=payload, verify=False)

                if response.status_code == 201:
                    messagebox.showinfo("Éxito", f"Monitor con Serial Number '{asset_data['serial']}' registrado correctamente en GLPI.")
                else:
                    messagebox.showerror("Error", f"Error al registrar el monitor en GLPI: {response.status_code}")
                    try:
                        messagebox.showerror("Error", response.json())
                    except json.JSONDecodeError:
                        messagebox.showerror("Error", response.text)
        except Exception as e:
            messagebox.showerror("Error", f"Error al subir el monitor a GLPI: {str(e)}")

    def procesar_qr_monitor(self, qr_data):
        """
        Genera la plantilla de datos para un monitor basado en su número de serie.
        """
        try:
            # Obtener sesión de GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión en GLPI.")
                return None

            # Solicitar ubicación y obtener location_id
            location_name = simpledialog.askstring("Input", "Ingrese la ubicación del monitor:").strip()
            location_id = self.obtener_location_id(session_token, location_name)

            if not location_id:
                messagebox.showerror("Error", f"No se encontró la ubicación '{location_name}' en GLPI.")
                return None

            # Crear la plantilla de datos del monitor
            plantilla_monitor = {
                "asset_type": "Monitor",
                "name": f"Monitor-{qr_data}",
                "locations_id": location_id,
                "manufacturers_id": "Dell Inc.",
                "serial": qr_data.strip(),
                "status": "Stocked",
            }

            return plantilla_monitor

        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el monitor: {str(e)}")
            return None

    def registrar_monitor_en_excel(self, serial_number, manufacturer, location_name):
        try:
            # Crear diccionario con los datos del monitor
            asset_data = {
                "serial": serial_number,
                "manufacturers_id": manufacturer,
                "name": f"Monitor-{serial_number}",
                "status": "Stocked",
                "locations_id": location_name
            }
            
            # Agregar al Excel usando la función modularizada
            self.agregar_a_excel(asset_data, "Monitor New")
            messagebox.showinfo("Éxito", "Monitor registrado exitosamente en el Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar el monitor en Excel: {str(e)}")

    def registrar_monitor_en_glpi(self, serial_number, manufacturer, location_name):
        try:
            # Obtener sesión de GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            # Verificar si el asset ya existe en GLPI
            if not self.verificar_existencia_asset(session_token, serial_number, asset_type="Monitor"):
                # Solicitar ubicación y obtener location_id
                location_id = self.obtener_location_id(session_token, location_name)
                if not location_id:
                    messagebox.showerror("Error", f"No se encontró la ubicación '{location_name}' en GLPI.")
                    return
                
                # Obtener el ID del fabricante en GLPI
                manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer)
                if not manufacturer_id:
                    messagebox.showerror("Error", f"No se pudo encontrar el fabricante '{manufacturer}' en GLPI.")
                    return

                # Definir los encabezados HTTP 
                headers = {
                    "Content-Type": "application/json",
                    "Session-Token": session_token,
                    "App-Token": APP_TOKEN
                }

                # Preparar datos para la creación en GLPI
                payload = {
                    "input": {
                        "name": f"Monitor-{serial_number}",
                        "serial": serial_number,
                        "manufacturers_id": int(manufacturer_id),
                        "locations_id": int(location_id),
                        "status": "Stocked"
                    }
                }

                response = requests.post(f"{GLPI_URL}/Monitor", headers=headers, json=payload, verify=False)

                if response.status_code == 201:
                    messagebox.showinfo("Éxito", f"Monitor con Serial Number '{serial_number}' registrado correctamente en GLPI.")
                else:
                    messagebox.showerror("Error", f"Error al registrar el monitor en GLPI: {response.status_code}")
                    try:
                        messagebox.showerror("Error", response.json())
                    except json.JSONDecodeError:
                        messagebox.showerror("Error", response.text)
            else:
                messagebox.showerror("Error", f"El monitor ya existe en el GLPI")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def registrar_monitor(self, qr_data, type_register):
        try:
            messagebox.showinfo("Información", "--- Registrar Monitor en Excel y GLPI ---")

            # Obtener la ubicación
            location_name = simpledialog.askstring("Input", "Ingrese la ubicación del monitor:").strip()
            if not location_name:
                messagebox.showerror("Error", "Debe ingresar una ubicación válida.")
                return

            # Seleccionar el fabricante (puedes modificar o agregar más fabricantes si es necesario)
            manufacturer = self.menu_emergente_n_botones("Seleccionar Fabricante", "Seleccione el fabricante del monitor:", ["Dell", "HP", "Samsung", "LG"])
            if not manufacturer:
                messagebox.showerror("Error", "Debe seleccionar un fabricante.")
                return

            if type_register == "Excel":
                # Registrar en Excel
                self.registrar_monitor_en_excel(qr_data.strip(), manufacturer, location_name)
            elif type_register == "GLPI":
                # Registrar en GLPI
                self.registrar_monitor_en_glpi(qr_data.strip(), manufacturer, location_name)

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def entregar_monitor(self):
        messagebox.showinfo("Información", "--- Entregar Monitor a Usuario ---")
        #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el número de serie manualmente? (escanear/manual):").strip().lower()
        metodo = self.menu_emergente_botones("Input", f"¿Desea escanear el QR o ingresar el número de serie manualmente?:", "Escanear", "Manual")

        if metodo == "Escanear":
            qr_data = self.escanear_qr_con_celular()
            if re.match(r'^CN[A-Z0-9]{10}$', qr_data) or re.match(r'^S?[A-Z0-9]{7}$', qr_data) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', qr_data):
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")
                serial_number = qr_data.strip()
            else:
                messagebox.showerror("Error", "Código QR no corresponde a un monitor válido.")
                return
        elif metodo == "Manual":
            res = simpledialog.askstring("Input", "Ingrese el número de serie del monitor:").strip()
            if re.match(r'^CN[A-Z0-9]{10}$', res) or re.match(r'^S?[A-Z0-9]{7}$', res) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', res):
                serial_number = res
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")
            else:
                messagebox.showerror("Error", "Código QR no corresponde a un monitor válido.")
                return
                
        else:
            messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
            return

        df = pd.read_excel(ruta_excel, sheet_name="Monitor")
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        filtro = df[df["serial"].str.lower() == serial_number.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró un monitor con el número de serie '{serial_number}' en el archivo Excel.")
            return

        nuevo_usuario = simpledialog.askstring("Input", "Ingrese el nombre del usuario que recibirá el monitor:").strip()

        if not nuevo_usuario:
            messagebox.showerror("Error", "El nombre del usuario no puede estar vacío.")
            return

        # Convertir a string
        nuevo_usuario = str(nuevo_usuario).strip()

        manufacturer = "Dell Inc."

        # Manejar valores NaN antes de actualizar el DataFrame
        df["users_id"] = df["users_id"].fillna("")
        df["name"] = df["name"].fillna("Unknown")

        # Determinar el nuevo nombre del monitor en base al usuario
        #manufacturer = filtro["manufacturers_id"].values[0]
        if manufacturer == "Dell" or manufacturer == "dell" or manufacturer == "Dell inc." or manufacturer == "Dell Inc." or manufacturer == "DELL":
            new_name = f"{nuevo_usuario}-DellMonitor"
        #elif manufacturer == "Samsung":
        #    new_name = f"{nuevo_usuario}-SamsungMonitor"
        #else:
        #    new_name = f"{nuevo_usuario}-Monitor"

        df.loc[df["serial"].str.lower() == serial_number.lower(), "users_id"] = nuevo_usuario
        df.loc[df["serial"].str.lower() == serial_number.lower(), "name"] = new_name
        df.loc[df["serial"].str.lower() == serial_number.lower(), "manufacturers_id"] = manufacturer

        df.to_excel(ruta_excel, sheet_name="Monitor", index=False)
        messagebox.showinfo("Información", f"Monitor con número de serie '{serial_number}' asignado a '{nuevo_usuario}' en el Excel.")

        # Actualizar en GLPI

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        usuario_id = self.obtener_id_usuario(session_token, nuevo_usuario)
        if not usuario_id:
            messagebox.showerror("Error", f"No se encontró el usuario '{nuevo_usuario}' en GLPI. 2")
            return
        #else:
            #messagebox.showinfo("Información", f"Se encontró el usuario '{nuevo_usuario}' en GLPI con el ID {usuario_id}.")

        # Obtener el ID del fabricante en GLPI
        manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer)
        if not manufacturer_id:
            messagebox.showerror("Error", f"No se encontró el fabricante '{manufacturer}' en GLPI.")
            return
        #else:
            #messagebox.showinfo("Información", f"Se encontró el fabricante '{manufacturer}' en GLPI con el ID {manufacturer_id}.")

        asset_id = self.obtener_asset_id_por_serial_monitor(session_token, serial_number)
        if not asset_id:
            messagebox.showerror("Error", "No se pudo encontrar el activo en GLPI.")
            return

        asset_data = filtro.iloc[0].to_dict()
        asset_data["users_id"] = nuevo_usuario
        asset_data["name"] = new_name 
        asset_data["manufacturers_id"] = manufacturer

        self.actualizar_asset_glpi_monitor(session_token, asset_id, asset_data)

    def actualizar_asset_glpi_monitor(self, session_token, asset_id, asset_data):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        user_name = str(asset_data["users_id"]).strip()

        # Obtener el ID del usuario basado en su nombre
        user_id = self.obtener_id_usuario(session_token, user_name)

        # Determinar el nuevo nombre del monitor
        fabricante = asset_data["manufacturers_id"]
        if "Dell Inc." == fabricante:
            new_name = f"{asset_data['users_id']}-DellMonitor"
        elif "Samsung" == fabricante:
            new_name = f"{asset_data['users_id']}-SamsungMonitor"
        else:
            new_name = f"{asset_data['users_id']}-Monitor"

        manufacturer_id = self.obtener_manufacturer_id(session_token, fabricante)
        # Preparar datos para la actualización en GLPI
        payload = {
            "input": {
                "id": asset_id,  
                "name": new_name,
                "manufacturer_id": manufacturer_id,
                "users_id": user_id
            }
        }

        response = requests.put(f"{GLPI_URL}/Monitor/{asset_id}", headers=headers, json=payload, verify=False)

        if response.status_code == 200:
            messagebox.showinfo("Éxito", f"Monitor con ID {asset_id} actualizado correctamente en GLPI con el nombre '{new_name}'.")
        else:
            messagebox.showerror("Error", f"Error al actualizar el monitor en GLPI: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def obtener_id_por_nombre_monitor(self, session_token, asset_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": asset_name.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/Monitor", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            for asset in response.json():
                if asset.get("name").strip().lower() == asset_name.strip().lower():
                    messagebox.showinfo("Información", f"Monitor encontrado: {asset_name}, ID: {asset.get('id')}")
                    return asset.get("id")

            messagebox.showinfo("Información", f"No se encontró el ID para el monitor '{asset_name}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el ID del monitor: {response.status_code}")
            return None

    def obtener_asset_id_por_serial_monitor(self, session_token, serial_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": serial_number.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/search/Monitor", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            assets = response.json().get("data", [])
            
            for asset in assets:
                serial_found = (asset.get("5") or "").strip().lower()  # Clave 5 es el serial number
                asset_name = asset.get("1")  # Clave 1 es el nombre del asset
                
                if serial_found == serial_number.lower():
                    messagebox.showinfo("Información", f"Monitor encontrado: {asset_name}, Serial: {serial_number}")
                    return self.obtener_id_por_nombre_monitor(session_token, asset_name)

            messagebox.showinfo("Información", f"No se encontró un monitor con el número de serie '{serial_number}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el monitor en GLPI: {response.status_code}")
            return None

    ## ----- Laptops -------

    def procesar_qr_laptop(self, flag, qr_data):
        """Genera la estructura de datos para registrar una laptop en Excel y GLPI."""

        # Diccionario con plantillas para cada fabricante
        plantillas = {
            "Dell": {
                "manufacturers_id": "Dell Inc.",
                "model": "Latitude"
            },
            "Mac": {
                "manufacturers_id": "Apple Inc.",
                "model": "MacBook Pro"
            }
        }

        if flag not in plantillas:
            messagebox.showerror("Error", "Fabricante no reconocido.")
            return None

        # Crear la plantilla base
        plantilla = {
            "asset_type": "Computer",
            "name": None,  # Se generará dinámicamente
            "locations_id": None,  # Solicitar al usuario
            "manufacturers_id": plantillas[flag]["manufacturers_id"],
            "serial": qr_data.strip(),  # QR escaneado de la laptop
            "computertypes_id": "Laptop",
            "status": "Stocked",  # Estado inicial
            "location": None,
            "model": plantillas[flag]["model"]
        }

        # Generar el nombre dinámicamente (opcional, si se requiere usuario más adelante)
        plantilla["name"] = f"{plantilla['manufacturers_id']}-{plantilla['model']}"

        return plantilla

    def manejar_qr_laptop(self, flag):
        try: 
            if flag == "Register":
                #manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                manufacturer = self.menu_emergente_botones("Input", "Ingrese el fabricante del laptop (Dell/Mac):", "dell", "mac")
                serial_number = None
                
                if manufacturer in ["dell", "dell inc.", "dell inc", "dell inc.", "dell"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Dell Inc."
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Dell Inc."
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                elif manufacturer in ["mac", "mac inc.", "apple inc.", "apple"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{9}$', qr_data) or re.match(r'^S[A-Za-z0-9]{12}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Apple Inc"
                                    return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido o no corresponde al manufacturer. Intente nuevamente.")
                            return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    #messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    #asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    #self.agregar_a_excel(asset_data, "Computer")
                                    manufacturer = "Apple Inc"
                                    return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido no corresponde al manufacturer. Intente nuevamente.")
                            return
                else: 
                    messagebox.showerror("Error", "Fabricante no válido. Intente nuevamente.")
                    return
            elif flag == "Deliver":
                #manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                manufacturer = self.menu_emergente_botones("Input", "Ingrese el fabricante del laptop (Dell/Mac):", "dell", "mac")
                serial_number = None
                
                if manufacturer in ["dell", "dell inc.", "dell inc", "dell inc.", "dell"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Dell Inc."
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Dell Inc."
                                return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                    
                elif manufacturer in ["mac", "mac inc.", "apple inc.", "apple"]:
                    #metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    metodo = self.seleccionar_metodo_ingreso()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Apple Inc"
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            #confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            confirmacion = self.menu_emergente_botones("Confirmación", f"¿Es correcto este Serial Number: {serial_number}, desea continuar?:", "Si", "No")
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                manufacturer = "Apple Inc"
                                return serial_number, manufacturer 
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return    
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                        
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
            return
            
    def registrar_laptop_en_excel(self, serial_number, manufacturer, model, location_name):
        try:
            # Crear diccionario con los datos del laptop
            asset_data = {
                "serial": serial_number,
                "manufacturers_id": manufacturer,
                "name": f"{manufacturer}-{model}",
                "status": "Stocked",  # Estado inicial en inventario
                "locations_id": location_name,  # ID de ubicación obtenido de GLPI
                "computermodels_id": model
            }
            
            # Agregar al Excel usando la función modularizada
            self.agregar_a_excel(asset_data, "Computer New")
            messagebox.showinfo("Éxito", "Laptop registrada exitosamente en el Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar la laptop en Excel: {str(e)}")

    def registrar_laptop_en_glpi(self, serial_number, manufacturer, model, location_name):
        try:
            # Obtener sesión de GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            # Verificar si el asset ya existe en GLPI
            if not self.verificar_existencia_asset(session_token, serial_number, asset_type="Computer"):
                # Solicitar ubicación y obtener location_id
                location_id = self.obtener_location_id(session_token, location_name)
                if not location_id:
                    messagebox.showerror("Error", f"No se encontró la ubicación '{location_name}' en GLPI.")
                    return
                
                # Obtener el ID del fabricante en GLPI
                manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer)
                if not manufacturer_id:
                    messagebox.showerror("Error", f"No se pudo encontrar el fabricante '{manufacturer}' en GLPI.")
                    return

                # Definir los encabezados HTTP 
                headers = {
                    "Content-Type": "application/json",
                    "Session-Token": session_token,
                    "App-Token": APP_TOKEN
                }

                # Preparar datos para la creación en GLPI
                payload = {
                    "input": {
                        "name": f"{manufacturer}-{model}",
                        "serial": serial_number,
                        "manufacturers_id": int(manufacturer_id),
                        "locations_id": int(location_id),  # Se agrega el location_id a GLPI
                        "status": "Stocked"
                    }
                }

                response = requests.post(f"{GLPI_URL}/Computer", headers=headers, json=payload, verify=False)

                if response.status_code == 201:
                    messagebox.showinfo("Éxito", f"Laptop con Service Tag '{serial_number}' registrada correctamente en GLPI.")
                else:
                    messagebox.showerror("Error", f"Error al registrar el laptop en GLPI: {response.status_code}")
                    try:
                        messagebox.showerror("Error", response.json())
                    except json.JSONDecodeError:
                        messagebox.showerror("Error", response.text)
            else:
                messagebox.showerror("Error", f"El laptop ya existe en el GLPI")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def registrar_laptop(self, type_register):
        try:
            messagebox.showinfo("Información", "--- Registrar Laptop en Excel y GLPI ---")
            
            # Manejar el QR para obtener el serial y fabricante
            result = self.manejar_qr_laptop("Register")
            if result is None:
                messagebox.showerror("Error", "No se pudo obtener el serial number y el fabricante del laptop.")
                return
            
            serial_number, manufacturer = result
        
            # Determinar el nuevo nombre del laptop según el fabricante
            if manufacturer == "Dell Inc.":
                models_dict = self.buscar_modelos_latitude_precision(self.obtener_token_sesion())
                models_lista = list(models_dict.values())
                
                model = self.menu_emergente_n_botones("Input Modelo Dell", "Ingrese el modelo Dell (Latitude/Precision):", models_lista)
                if model == "Latitude":
                    new_name = "None-Latitude"
                elif model == "Precision":
                    new_name = "None-Precision"
                else: 
                    messagebox.showerror("Error", "No se introdujo el modelo correctamente.")
            elif manufacturer == "Apple Inc": 
                model = "macbook"
                new_name = "None-MacBookPro"
            else:
                messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
                return
            
            location_name = simpledialog.askstring("Input", "Ingrese la ubicación del activo:").strip()

            if type_register == "Excel":
                # Registrar en Excel
                self.registrar_laptop_en_excel(serial_number, manufacturer, model, location_name)
            elif type_register == "GLPI":
                # Registrar en GLPI
                self.registrar_laptop_en_glpi(serial_number, manufacturer, model, location_name)

        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def obtener_location_id(self, session_token, location_name):
        if pd.isna(location_name) or location_name is None:
            messagebox.showerror("Error", "Ubicación no proporcionada. No se puede registrar el consumible.")
            return None

        location_name = str(location_name).strip()  # Convertir a string y limpiar espacios

        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        params = {'searchText': location_name, 'range': '0-999'}
        response = requests.get(f"{GLPI_URL}/Location", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            locations = response.json()
            for location in locations:
                if location.get("name", "").strip().lower() == location_name.lower():
                    return location["id"]
        
        messagebox.showerror("Error", f"No se encontró la ubicación '{location_name}' en GLPI.")
        return None

    def entregar_laptop(self):
        try: 
            messagebox.showinfo("Información", "--- Entregar Laptop a Usuario ---")
            result = self.manejar_qr_laptop("Deliver")
            
            if not result:
                messagebox.showerror("Error", "No se pudo obtener el serial number y el fabricante del laptop.")
                return
            
            serial_number, manufacturer = result

            # Cargar el archivo Excel
            wb = load_workbook(ruta_excel)
            if "Computer" not in wb.sheetnames:
                messagebox.showerror("Error", "La hoja 'Computer' no existe en el archivo Excel.")
                return
            
            ws = wb["Computer"]
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]
            df = df[1:]

            if df.empty:
                messagebox.showerror("Error", "El archivo Excel está vacío.")
                return

            # Validar columnas necesarias
            required_columns = ["serial", "manufacturers_id", "users_id", "name"]
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("Error", "El archivo Excel no contiene las columnas necesarias.")
                return
            
            filtro = df[df["serial"].str.lower() == serial_number.lower()]

            if filtro.empty:
                messagebox.showerror("Error", f"No se encontró un laptop con el Service Tag '{serial_number}' en el archivo Excel.")
                return

            nuevo_usuario = simpledialog.askstring("Input", "Ingrese el nombre del usuario que recibirá el laptop:")
            if not nuevo_usuario:
                messagebox.showerror("Error", "El nombre del usuario no puede estar vacío.")
                return
            
            # Convertir a string
            nuevo_usuario = str(nuevo_usuario).strip()

            # Obtener el ID del usuario en GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            usuario_id = self.obtener_id_usuario(session_token, nuevo_usuario)
            if not usuario_id:
                messagebox.showerror("Error", f"No se encontró el usuario '{nuevo_usuario}' en GLPI. 2")
                return
            else:
                messagebox.showinfo("Información", f"Se encontró el usuario '{nuevo_usuario}' en GLPI con el ID {usuario_id}.")
            
            # Obtener el ID del fabricante en GLPI
            manufacturer_id = self.obtener_manufacturer_id(session_token, manufacturer)
            if not manufacturer_id:
                messagebox.showerror("Error", f"No se encontró el fabricante '{manufacturer}' en GLPI.")
                return
            else:
                messagebox.showinfo("Información", f"Se encontró el fabricante '{manufacturer}' en GLPI con el ID {manufacturer_id}.")

            # Asegurar que la columna 'users_id' sea string antes de modificarla
            df["users_id"] = df["users_id"].astype(str).fillna("")
            df["name"] = df["name"].fillna("Unknown")

            # Determinar el nuevo nombre del laptop en base al fabricante
            if manufacturer.lower() in ["dell", "dell inc.", "Dell Inc."]:
                new_name = f"{nuevo_usuario}-Latitude"
            elif manufacturer.lower() in ["apple", "apple inc.", "mac", "Apple Inc"]:
                new_name = f"{nuevo_usuario}-MacBookPro"
            else:
                messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
                return

            # Actualizar DataFrame con los nuevos valores
            df.loc[df["serial"].str.lower() == serial_number.lower(), "users_id"] = usuario_id
            df.loc[df["serial"].str.lower() == serial_number.lower(), "name"] = new_name
            df.loc[df["serial"].str.lower() == serial_number.lower(), "manufacturers_id"] = manufacturer_id

            # Limpiar las filas de datos existentes antes de escribir los datos actualizados
            ws.delete_rows(2, ws.max_row)

            # Escribir los datos actualizados de vuelta a la hoja "Computer" sin los encabezados
            for row in dataframe_to_rows(df, index=False, header=False):
                ws.append(row)

            wb.save(ruta_excel)
            messagebox.showinfo("Información", f"Laptop con Service Tag '{serial_number}' asignado a '{nuevo_usuario}' en el Excel.")

            # Actualizar en GLPI
            asset_id = self.obtener_asset_id_por_serial(session_token, serial_number)
            if not asset_id:
                messagebox.showerror("Error", "No se pudo encontrar el activo en GLPI.")
                return

            asset_data = filtro.iloc[0].to_dict()
            asset_data["users_id"] = nuevo_usuario
            asset_data["name"] = new_name
            asset_data["manufacturers_id"] = manufacturer  # Ahora se pasa el ID del fabricante

            self.actualizar_asset_glpi(session_token, asset_id, asset_data)
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def obtener_asset_id_por_serial(self, session_token, serial_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": serial_number.strip().lower(),
            "range": "0-999"
        }

        # Buscar el asset por número de serie
        response = requests.get(f"{GLPI_URL}/search/Computer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            assets = response.json().get("data", [])
            
            for asset in assets:
                serial_found = (asset.get("5") or "").strip().lower()  # Clave 5 es el serial number
                asset_name = asset.get("1")  # Clave 1 es el nombre del asset
                
                if serial_found == serial_number.lower():
                    messagebox.showinfo("Información", f"Activo encontrado: {asset_name}, Serial: {serial_number}")
                    # Ahora buscamos el ID utilizando el nombre del activo encontrado
                    return self.obtener_id_por_nombre(session_token, asset_name)

            messagebox.showinfo("Información", f"No se encontró un activo con el serial number '{serial_number}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el activo en GLPI: {response.status_code}")
            return None

    def obtener_id_por_nombre(self, session_token, asset_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": asset_name.strip().lower(),
            "range": "0-999"
        }

        # Buscar el asset por nombre
        response = requests.get(f"{GLPI_URL}/Computer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            for asset in response.json():
                if asset.get("name").strip().lower() == asset_name.strip().lower():
                    messagebox.showinfo("Información", f"Activo encontrado: {asset_name}, ID: {asset.get('id')}")
                    return asset.get("id")

            messagebox.showinfo("Información", f"No se encontró el ID para el activo '{asset_name}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el ID del activo: {response.status_code}")
            return None
    
    def actualizar_asset_glpi(self, session_token, asset_id, asset_data):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Asegurar que users_id es un string antes de buscarlo en GLPI
        
        #user_name = str(asset_data["users_id"]).strip()
        
        user_name = str(asset_data["users_id"]).strip()
        user_id = self.obtener_id_usuario(session_token, user_name)
        resultado = self.obtener_name_usuario(session_token, user_name)

        if not resultado:
            messagebox.showerror("Error", f"No se encontró el usuario '{user_name}' en GLPI. 6")
            return  # Evita que el código continúe con None

        name_full, name_full_glpi = resultado

        #print(name_full)
        #print(name_full_glpi)
        messagebox.showinfo("Información", f"ID del usuario {user_name} encontrado y listo para GLPI")
        
        if not user_name:
            messagebox.showerror("Error", f"No se encontró el usuario '{user_name}' en GLPI. 5")
            return

        # Determinar el nuevo nombre según el fabricante
        fabricante = asset_data["manufacturers_id"]
        if "Dell Inc." == fabricante:
            new_name = f"{user_name}-Latitude"
        elif "Apple Inc" == fabricante:
            new_name = f"{user_name}-MacBookPro"
        else:
            messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
            return

        manufacturer_id = self.obtener_manufacturer_id(session_token, fabricante)

        # Preparar datos para la actualización en GLPI
        payload = {
            "input": {
                "id": asset_id,  
                "name": new_name,
                "manufacturer_id": manufacturer_id,
                "users_id": user_id
            }
        }

        response = requests.put(f"{GLPI_URL}/Computer/{asset_id}", headers=headers, json=payload, verify=False)

        if response.status_code == 200:
            messagebox.showinfo("Éxito", f"Activo con ID {asset_id} actualizado correctamente en GLPI con el nombre '{new_name}'.")
        else:
            messagebox.showerror("Error", f"Error al actualizar el activo en GLPI: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def obtener_id_usuario(self, session_token, nombre_usuario):
        """Obtiene el ID de un usuario en GLPI a partir de su nombre"""
        headers = {
            "Session-Token": session_token,
            "App-Token": APP_TOKEN,
            "Content-Type": "application/json"
        }
        
        params = {"searchText": str(nombre_usuario).strip(), "range": "0-999"}  # Asegurar que sea string
        response = requests.get(f"{GLPI_URL}/User", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            usuarios = response.json()
            for usuario in usuarios:
                # Evitar errores con valores None o numéricos
                apellido = str(usuario.get("realname", "")).strip()
                nombre = str(usuario.get("firstname", "")).strip()

                nombre_completo_glpi = f"{apellido} {nombre}".strip()
                nombre_completo = f"{nombre} {apellido}".strip()

                if str(usuario.get("name", "")).strip().lower() == nombre_usuario.strip().lower() or nombre_completo_glpi.lower() == nombre_usuario.strip().lower() or nombre_completo.lower() == nombre_usuario.strip().lower():
                    return usuario["id"]

            print(f"No se encontró el usuario '{nombre_usuario}' en GLPI. 4")
            return None
        else:
            print(f"Error al buscar usuario '{nombre_usuario}': {response.status_code}")
            return None

    def obtener_name_usuario(self, session_token, nombre_usuario):
        """Obtiene el nombre de un usuario en GLPI a partir de su búsqueda"""
        headers = {
            "Session-Token": session_token,
            "App-Token": APP_TOKEN,
            "Content-Type": "application/json"
        }
        
        params = {"searchText": str(nombre_usuario).strip(), "range": "0-999"}
        response = requests.get(f"{GLPI_URL}/User", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            usuarios = response.json()
            for usuario in usuarios:
                apellido = str(usuario.get("realname", "")).strip()
                nombre = str(usuario.get("firstname", "")).strip()

                nombre_completo_glpi = f"{apellido} {nombre}".strip()
                nombre_completo = f"{nombre} {apellido}".strip()

                if str(usuario.get("name", "")).strip().lower() == nombre_usuario.strip().lower() or \
                nombre_completo_glpi.lower() == nombre_usuario.strip().lower() or \
                nombre_completo.lower() == nombre_usuario.strip().lower():
                    return nombre_completo, nombre_completo_glpi

            print(f"No se encontró el usuario '{nombre_usuario}' en GLPI. 3")
            return "", ""  # Devolver una tupla vacía en lugar de None
        else:
            print(f"Error al buscar usuario '{nombre_usuario}': {response.status_code}")
            return "", ""  # Devolver una tupla vacía en caso de error

    def obtener_locations(self, session_token):
        try:
            headers = {
                "Content-Type": "application/json",
                "Session-Token": session_token,
                "App-Token": APP_TOKEN
            }

            params = {"range": "0-999"}

            response = requests.get(f"{GLPI_URL}/Location", headers=headers, params=params, verify=False)

            if response.status_code == 200:
                locations = response.json()

                # Construir un diccionario {id: name}
                location_dict = {location["id"]: location["name"] for location in locations if "id" in location and "name" in location}

                #print(location_dict)

            else:
                print(f"No se pudo obtener la lista de ubicaciones. Código: {response.status_code}")
                return {}

        except Exception as e:
            print(f"Error al obtener ubicaciones: {str(e)}")
            return {}

    def buscar_modelos_glpi(self, session_token):
        """
        Usa el endpoint de búsqueda de GLPI para encontrar modelos de computadoras.
        """
        url = f"{GLPI_URL}/ComputerModel"
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        params = {"range": "0-999"}

        response = requests.get(url, headers=headers, params=params, verify=False)

        if response.status_code == 200:
            models = response.json()
            models_dict = {model["id"]: model["name"] for model in models if "id" in model and "name" in model}
            #print(models_dict)
            return(models_dict)
        else:
            messagebox.showerror("Error", f"No se pudo buscar modelos. Código: {response.status_code}\n{response.text}")
            return None

    def buscar_modelos_latitude_precision(self, session_token):
        """
        Usa el endpoint de búsqueda de GLPI para encontrar modelos de computadoras.
        Filtra solo los modelos 'Latitude {número}' o 'Precision {número}'.
        """
        url = f"{GLPI_URL}/ComputerModel"
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        params = {"range": "0-999"}

        response = requests.get(url, headers=headers, params=params, verify=False)

        if response.status_code == 200:
            models = response.json()
            
            # Expresión regular para detectar "Latitude {número}" o "Precision {número}"
            regex = re.compile(r"^(Latitude|Precision) \d+$")

            # Filtrar modelos que cumplan con la expresión regular
            filtered_models = {model["id"]: model["name"] for model in models 
                            if "id" in model and "name" in model and regex.match(model["name"])}

            #print(filtered_models)  # Mostrar en consola para verificar
            return filtered_models  # Retorna solo los modelos Latitude o Precision con números

        else:
            messagebox.showerror("Error", f"No se pudo buscar modelos. Código: {response.status_code}\n{response.text}")
            return None

    def obtener_modelo_id(self, session_token, modelo_nombre):
        """
        Obtiene el ID de un modelo de computadora en GLPI a partir de su nombre.
        """
        modelos_dict = self.buscar_modelos_glpi(session_token)  # Obtener todos los modelos

        if modelos_dict is None:
            messagebox.showerror("Error", "No se pudieron obtener modelos de GLPI.")
            return None

        # Buscar el modelo por nombre (ignorando mayúsculas y espacios)
        modelo_id = None
        for id_, nombre in modelos_dict.items():
            if nombre.strip().lower() == modelo_nombre.strip().lower():
                modelo_id = id_
                break

        if modelo_id:
            return modelo_id
        else:
            messagebox.showerror("Error", f"No se encontró el modelo '{modelo_nombre}' en GLPI.")
            return None

    def obtener_modelo_dell(service_tag):
        url = f"https://api.dell.com/support/v4/assetinfo/{service_tag}"
        headers = {
            "Accept": "application/json",
            "Authorization": "Bearer TU_ACCESS_TOKEN"  # Necesitas obtener un API Key de Dell
        }

        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            modelo = data.get("Model", "Modelo no encontrado")
            return modelo
        else:
            print(f"Error al obtener datos de Dell: {response.status_code}")
            return None


if __name__ == "__main__":
    root = tk.Tk()
    app = GLPIApp(root)
    root.mainloop()